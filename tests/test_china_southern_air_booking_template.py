import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.services.china_southern_air_booking_excel import (
    ChinaSouthernAirBookingExcelError,
    ChinaSouthernAirBookingExcelService,
)


class ChinaSouthernAirBookingTemplateTests(unittest.TestCase):
    TEMPLATE_PATH = (
        Path(__file__).resolve().parents[1]
        / "documents"
        / "china_southern_air"
        / "南航订舱模板.xlsx"
    )

    def test_adds_dynamic_cargo_type_dropdown_without_changing_visible_template(self):
        labels = ["普货", "贵重物品", "活体动物"]
        content = ChinaSouthernAirBookingExcelService.build_template(
            self.TEMPLATE_PATH,
            cargo_type_labels=labels,
        )

        source = load_workbook(self.TEMPLATE_PATH)
        generated = load_workbook(BytesIO(content))
        try:
            source_sheet = source.worksheets[0]
            generated_sheet = generated.worksheets[0]

            self.assertEqual(source_sheet.title, generated_sheet.title)
            self.assertEqual(source_sheet.max_row, generated_sheet.max_row)
            self.assertEqual(source_sheet.max_column, generated_sheet.max_column)
            self.assertEqual(
                list(source_sheet.values),
                list(generated_sheet.values),
            )
            self.assertEqual(
                {str(item) for item in source_sheet.merged_cells.ranges},
                {str(item) for item in generated_sheet.merged_cells.ranges},
            )
            self.assertEqual(
                [
                    worksheet.title
                    for worksheet in generated.worksheets
                    if worksheet.sheet_state == "visible"
                ],
                [source_sheet.title],
            )
            self.assertEqual(
                source_sheet.sheet_view.showGridLines,
                generated_sheet.sheet_view.showGridLines,
            )
            for row_number, source_dimension in source_sheet.row_dimensions.items():
                self.assertEqual(
                    source_dimension.height,
                    generated_sheet.row_dimensions[row_number].height,
                )
            for column_name, source_dimension in source_sheet.column_dimensions.items():
                self.assertEqual(
                    source_dimension.width,
                    generated_sheet.column_dimensions[column_name].width,
                )
            for row in source_sheet.iter_rows():
                for source_cell in row:
                    generated_cell = generated_sheet[source_cell.coordinate]
                    self.assertEqual(source_cell._style, generated_cell._style)

            source_validations = {
                (str(item.sqref), item.formula1)
                for item in source_sheet.data_validations.dataValidation
            }
            generated_validations = {
                (str(item.sqref), item.formula1)
                for item in generated_sheet.data_validations.dataValidation
            }
            self.assertTrue(source_validations.issubset(generated_validations))
            self.assertIn(
                ("F4:F24", ChinaSouthernAirBookingExcelService.CARGO_TYPE_OPTIONS_RANGE),
                generated_validations,
            )
            cargo_type_validation = next(
                item
                for item in generated_sheet.data_validations.dataValidation
                if str(item.sqref) == "F4:F24"
            )
            self.assertFalse(cargo_type_validation.showInputMessage)
            self.assertIsNone(cargo_type_validation.promptTitle)
            self.assertIsNone(cargo_type_validation.prompt)
            self.assertTrue(cargo_type_validation.showErrorMessage)

            options_sheet = generated[
                ChinaSouthernAirBookingExcelService.CARGO_TYPE_OPTIONS_SHEET
            ]
            self.assertEqual(options_sheet.sheet_state, "veryHidden")
            self.assertEqual(
                [options_sheet.cell(row=index, column=1).value for index in range(1, 4)],
                labels,
            )
            defined_name = generated.defined_names[
                ChinaSouthernAirBookingExcelService.CARGO_TYPE_OPTIONS_RANGE
            ]
            self.assertEqual(
                defined_name.attr_text,
                "'_cargo_type_options'!$A$1:$A$3",
            )
        finally:
            source.close()
            generated.close()

    def test_uses_named_range_when_labels_exceed_inline_list_limit(self):
        labels = [f"动态货物类型{index:03d}" for index in range(1, 101)]
        self.assertGreater(len(",".join(labels)), 255)

        content = ChinaSouthernAirBookingExcelService.build_template(
            self.TEMPLATE_PATH,
            cargo_type_labels=labels,
        )
        generated = load_workbook(BytesIO(content))
        try:
            validation = next(
                item
                for item in generated.worksheets[0].data_validations.dataValidation
                if str(item.sqref) == "F4:F24"
            )
            self.assertEqual(
                validation.formula1,
                ChinaSouthernAirBookingExcelService.CARGO_TYPE_OPTIONS_RANGE,
            )
            self.assertEqual(
                generated[ChinaSouthernAirBookingExcelService.CARGO_TYPE_OPTIONS_SHEET][
                    "A100"
                ].value,
                labels[-1],
            )
        finally:
            generated.close()

    def test_rejects_empty_or_duplicate_dictionary_labels(self):
        for labels, message in (
            ([], "不存在或没有启用的选项"),
            (["普货", "普货"], "存在重复名称"),
            (["普货", " "], "存在空名称"),
        ):
            with self.subTest(labels=labels):
                with self.assertRaisesRegex(ChinaSouthernAirBookingExcelError, message):
                    ChinaSouthernAirBookingExcelService.build_template(
                        self.TEMPLATE_PATH,
                        cargo_type_labels=labels,
                    )


if __name__ == "__main__":
    unittest.main()
