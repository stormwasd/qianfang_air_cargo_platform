import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from app.services.cost_excel_export import (
    COST_EXPORT_HEADERS,
    append_cost_export_headers,
    format_bill_of_lading_for_export,
)


class CostExcelBillOfLadingTests(unittest.TestCase):
    def test_master_split_codes_are_converted(self):
        expected = {
            "1-0": "一主",
            "1-1": "一主（一）分",
            "1-2": "一主（二）分",
            "1-3": "一主（三）分",
            "1-4": "一主（四）分",
            "1-5": "一主（五）分",
            "1-6": "一主（六）分",
            "1-7": "一主（七）分",
            "1-8": "一主（八）分",
            "1-9": "一主（九）分",
        }
        for code, label in expected.items():
            with self.subTest(code=code):
                self.assertEqual(format_bill_of_lading_for_export(code), label)

    def test_direct_waybill_codes_are_converted(self):
        expected = {"2-0": "直单（虚拟分单）"}
        expected.update(
            {
                f"2-{index}": f"直单（虚拟分单*{index}）"
                for index in range(1, 10)
            }
        )
        for code, label in expected.items():
            with self.subTest(code=code):
                self.assertEqual(format_bill_of_lading_for_export(code), label)

    def test_current_stored_values_are_converted(self):
        expected = {
            "一主多分-0": "一主",
            "一主多分-1": "一主（一）分",
            "一主多分-6": "一主（六）分",
            "一主多分-9": "一主（九）分",
            "直单-0": "直单（虚拟分单）",
            "直单-1": "直单（虚拟分单*1）",
            "直单-3": "直单（虚拟分单*3）",
            "直单-9": "直单（虚拟分单*9）",
        }
        for stored_value, label in expected.items():
            with self.subTest(stored_value=stored_value):
                self.assertEqual(
                    format_bill_of_lading_for_export(stored_value),
                    label,
                )

    def test_empty_unknown_and_real_waybill_values_remain_compatible(self):
        self.assertEqual(format_bill_of_lading_for_export(None), "")
        self.assertEqual(format_bill_of_lading_for_export(""), "")
        self.assertEqual(format_bill_of_lading_for_export("3-1"), "3-1")
        self.assertEqual(
            format_bill_of_lading_for_export("784-98766543"),
            "784-98766543",
        )

    def test_both_export_layouts_write_converted_values_as_text(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.cell(row=2, column=7).value = format_bill_of_lading_for_export(
            "一主多分-6"
        )
        worksheet.cell(row=4, column=7).value = format_bill_of_lading_for_export(
            "直单-3"
        )

        self.assertEqual(worksheet.cell(row=2, column=7).value, "一主（六）分")
        self.assertEqual(worksheet.cell(row=2, column=7).data_type, "s")
        self.assertEqual(worksheet.cell(row=4, column=7).value, "直单（虚拟分单*3）")
        self.assertEqual(worksheet.cell(row=4, column=7).data_type, "s")

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "cost-export.xlsx"
            workbook.save(output_path)
            workbook.close()

            exported_workbook = load_workbook(output_path, read_only=True, data_only=True)
            exported_worksheet = exported_workbook.active
            self.assertEqual(exported_worksheet.cell(row=2, column=7).value, "一主（六）分")
            self.assertEqual(exported_worksheet.cell(row=2, column=7).data_type, "s")
            self.assertEqual(
                exported_worksheet.cell(row=4, column=7).value,
                "直单（虚拟分单*3）",
            )
            self.assertEqual(exported_worksheet.cell(row=4, column=7).data_type, "s")
            exported_workbook.close()


class CostExcelLayoutTests(unittest.TestCase):
    def test_removed_intl_air_columns_are_not_exported(self):
        self.assertEqual(len(COST_EXPORT_HEADERS), 115)
        self.assertNotIn("国空应付-航空公司", COST_EXPORT_HEADERS)
        self.assertNotIn("国空应付-托运日期", COST_EXPORT_HEADERS)
        # 国内空运属于另一业务分组，本次需求不应误删。
        self.assertIn("国空内应付-航空公司", COST_EXPORT_HEADERS)
        self.assertIn("汽运应付-托运日期", COST_EXPORT_HEADERS)
        self.assertIn("国空内应付-托运日期", COST_EXPORT_HEADERS)
        self.assertIn("地面应付-托运日期", COST_EXPORT_HEADERS)

    def test_export_titles_use_product_wording(self):
        expected_headers = {
            "应收-分单费 电报费/底账费",
            "应收-TC费",
            "应收-前置仓费",
            "国空应付-实际重量",
            "国空应付-单价",
            "国空应付-运费计算方式",
            "国空应付-燃油费",
            "国空应付-TC费",
            "国空内应付-实际重量",
            "国空内应付-运费计算方式",
        }
        for header in expected_headers:
            with self.subTest(header=header):
                self.assertIn(header, COST_EXPORT_HEADERS)

        old_headers = {
            "应收-分单费/抵账费/电报费",
            "应收-TC操作费/快件中心过站费",
            "应收-前置仓/国际货站地面费",
            "国空应付-重量",
            "国空应付-费率",
            "国空应付-借单/磁检/燃油/提货费",
            "国空应付-TC/入网/处置费",
            "国空内应付-重量",
        }
        for header in old_headers:
            with self.subTest(header=header):
                self.assertNotIn(header, COST_EXPORT_HEADERS)

    def test_grouped_headers_still_cover_all_columns(self):
        workbook = Workbook()
        worksheet = workbook.active

        headers = append_cost_export_headers(worksheet)

        merged_ranges = {str(item) for item in worksheet.merged_cells.ranges}
        self.assertEqual(len(headers), 115)
        self.assertEqual(worksheet.max_column, 115)
        self.assertIn("A1:Q2", merged_ranges)
        self.assertIn("AK1:DE1", merged_ranges)
        self.assertIn("DD2:DD3", merged_ranges)
        workbook.close()

    def test_every_export_section_keeps_its_expected_boundaries(self):
        expected_boundaries = {
            "应收-单价": 17,
            "应收-运费计算方式": 18,
            "应收-运费": 19,
            "国空应付-小计": 36,
            "国空应付-单价": 47,
            "国空应付-运费计算方式": 48,
            "国空应付-运费": 49,
            "国空应付-备注": 59,
            "汽运应付-小计": 60,
            "汽运应付-备注": 70,
            "国空内应付-小计": 71,
            "国空内应付-费率": 84,
            "国空内应付-运费计算方式": 85,
            "国空内应付-运费": 86,
            "国空内应付-备注": 88,
            "报关应付-小计": 89,
            "报关应付-备注": 96,
            "地面应付-小计": 97,
            "地面应付-备注": 107,
            "应付合计": 108,
            "利润率(%)": 114,
        }

        for header, expected_index in expected_boundaries.items():
            with self.subTest(header=header):
                self.assertEqual(COST_EXPORT_HEADERS.index(header), expected_index)


if __name__ == "__main__":
    unittest.main()
