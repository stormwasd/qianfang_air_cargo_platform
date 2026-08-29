"""南航批量订舱 Excel 解析服务。"""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile
from typing import Any, Dict, Iterable, List, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.exceptions import InvalidFileException


class ChinaSouthernAirBookingExcelError(Exception):
    """模板格式或行数据校验失败。"""

    def __init__(self, message: str, *, errors: Sequence[Dict[str, Any]] | None = None):
        super().__init__(message)
        self.errors = list(errors or [])


class ChinaSouthernAirBookingExcelService:
    """解析新版南航订舱模板，并构造平台订舱 ``form_data``。"""

    MAX_FILE_SIZE = 5 * 1024 * 1024
    MAX_BOOKING_ROWS = 500
    MAX_WORKSHEET_ROWS = 2000
    HEADER_SCAN_ROWS = 10
    HEADER_SCAN_COLUMNS = 60
    CARGO_TYPE_OPTIONS_SHEET = "_cargo_type_options"
    CARGO_TYPE_OPTIONS_RANGE = "CSA_CargoTypeOptions"

    HEADER_FIELDS = {
        "始发站": "origin_station",
        "到达站": "destination",
        "航班日期": "flight_date",
        "客户名称": "shipper_unit",
        "货物类型": "cargo_type",
        "货物代码": "cargo_code",
        "航班号": "flight_number",
        "宽体机订舱备注": "booking_remark_wide",
        "窄体机订舱备注": "booking_remark_narrow",
        "货物名称": "cargo_name",
        "件数": "quantity",
        "重量(kg)": "weight",
        "产品名称": "product_name",
        "超规货": "oversized_cargo",
        "特货码": "special_cargo_code",
        "订舱体积": "booking_volume",
        "储运注意事项": "storage_and_transportation_precautions",
        "无隐含危险品": "no_dangerous_goods",
        "出港货邮处理费选项": "outbound_cargo_and_mail_handling_fee_options",
    }
    HEADER_ALIASES = {
        "重量（kg）": "重量(kg)",
        "出港货邮处理费": "出港货邮处理费选项",
        "特货码（多个特货码用/隔开）": "特货码",
        "特货码(多个特货码用/隔开)": "特货码",
        "特货码（多个特货码用英文逗号隔开）": "特货码",
        "特货码(多个特货码用英文逗号隔开)": "特货码",
        "特货码（多个用英文逗号隔开）": "特货码",
        "特货码(多个用英文逗号隔开)": "特货码",
    }
    # 费用选项列保留在模板和映射中，但整列允许缺失；缺失时执行阶段
    # 直接透传南航费用查询返回的完整列表。
    REQUIRED_HEADERS = frozenset(HEADER_FIELDS) - {"出港货邮处理费选项"}
    REQUIRED_FIELDS = {
        "origin_station": "始发站",
        "destination": "到达站",
        "flight_date": "航班日期",
        "shipper_unit": "客户名称",
        "cargo_type": "货物类型",
        "cargo_code": "货物代码",
        "flight_number": "航班号",
        "cargo_name": "货物名称",
        "quantity": "件数",
        "weight": "重量(kg)",
    }
    # 这些列不包含模板预置值，可用于判断用户是否真正填写了该行。
    USER_INPUT_FIELDS = frozenset(
        {
            "destination",
            "flight_date",
            "shipper_unit",
            "cargo_type",
            "cargo_code",
            "flight_number",
            "booking_remark_wide",
            "booking_remark_narrow",
            "cargo_name",
            "quantity",
            "weight",
            "product_name",
            "special_cargo_code",
            "booking_volume",
            "storage_and_transportation_precautions",
            "outbound_cargo_and_mail_handling_fee_options",
        }
    )

    @staticmethod
    def _text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        if isinstance(value, bool):
            return "1" if value else "0"
        if isinstance(value, int):
            return str(value)
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return format(Decimal(str(value)).normalize(), "f")
        return str(value).strip()

    @classmethod
    def _normalized_header(cls, value: Any) -> str:
        text = cls._text(value).replace("\n", "").replace("\r", "").replace(" ", "")
        return cls.HEADER_ALIASES.get(text, text)

    @classmethod
    def _date_text(cls, value: Any, row_number: int) -> str:
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        text = cls._text(value)
        if not text:
            return ""
        normalized = text.replace("/", "-").replace(".", "-")
        try:
            return datetime.strptime(normalized, "%Y-%m-%d").date().isoformat()
        except ValueError as exc:
            raise ChinaSouthernAirBookingExcelError(
                f"第 {row_number} 行航班日期格式错误，应为 YYYY-MM-DD"
            ) from exc

    @classmethod
    def _flag_text(
        cls,
        value: Any,
        *,
        row_number: int,
        field_label: str,
        yes_value: str,
        no_value: str,
        default: str,
    ) -> str:
        text = cls._text(value)
        if not text:
            return default
        normalized = text.lower()
        if normalized in {"是", "yes", "y", "true"}:
            return yes_value
        if normalized in {"否", "no", "n", "false"}:
            return no_value
        if normalized in {"0", "1"}:
            return normalized
        raise ChinaSouthernAirBookingExcelError(
            f"第 {row_number} 行{field_label}只能填写“是”或“否”"
        )

    @classmethod
    def _find_header(cls, sheet) -> tuple[int, Dict[str, int]]:
        best_row = 0
        best_mapping: Dict[str, int] = {}
        if sheet.max_row is None or sheet.max_column is None:
            # 部分标准 xlsx（包括当前模板生成器）不写 worksheet dimension，
            # openpyxl 的只读模式需要先扫描一次才能得到实际范围。
            sheet.calculate_dimension(force=True)
        max_row = min(sheet.max_row, cls.HEADER_SCAN_ROWS)
        max_column = min(sheet.max_column, cls.HEADER_SCAN_COLUMNS)
        for row_number in range(1, max_row + 1):
            mapping: Dict[str, int] = {}
            for column_number in range(1, max_column + 1):
                header = cls._normalized_header(sheet.cell(row_number, column_number).value)
                if header in cls.HEADER_FIELDS:
                    mapping[header] = column_number
            if len(mapping) > len(best_mapping):
                best_row = row_number
                best_mapping = mapping

        missing = sorted(cls.REQUIRED_HEADERS - set(best_mapping))
        if missing:
            raise ChinaSouthernAirBookingExcelError(
                "Excel模板缺少必要列：" + "、".join(missing),
                errors=[{"type": "missing_headers", "headers": missing}],
            )
        return best_row, best_mapping

    @classmethod
    def build_template(
        cls,
        template_path: Path,
        *,
        cargo_type_labels: Iterable[str],
    ) -> bytes:
        """基于原模板动态添加货物类型下拉框，其他可见内容保持不变。"""
        labels: List[str] = []
        seen = set()
        for raw_label in cargo_type_labels:
            label = cls._text(raw_label)
            if not label:
                raise ChinaSouthernAirBookingExcelError(
                    "数据字典 nanfang_air_cargo_type 存在空名称的启用项"
                )
            if label in seen:
                raise ChinaSouthernAirBookingExcelError(
                    f"数据字典 nanfang_air_cargo_type 存在重复名称：{label}"
                )
            seen.add(label)
            labels.append(label)
        if not labels:
            raise ChinaSouthernAirBookingExcelError(
                "数据字典 nanfang_air_cargo_type 不存在或没有启用的选项"
            )

        try:
            workbook = load_workbook(template_path)
        except (BadZipFile, InvalidFileException, OSError, ValueError, KeyError, EOFError) as exc:
            raise ChinaSouthernAirBookingExcelError("南航订舱模板无法解析") from exc

        try:
            sheet = workbook.worksheets[0]
            header_row, header_mapping = cls._find_header(sheet)
            cargo_type_column = header_mapping["货物类型"]

            # 当前模板的表头纵向合并两行。根据合并区域计算真实填写起始行，
            # 避免把下拉校验误加到合并表头的占位单元格。
            header_end_row = header_row
            for merged_range in sheet.merged_cells.ranges:
                if (
                    merged_range.min_row <= header_row <= merged_range.max_row
                    and merged_range.min_col
                    <= cargo_type_column
                    <= merged_range.max_col
                ):
                    header_end_row = max(header_end_row, merged_range.max_row)
            data_start_row = header_end_row + 1
            data_end_row = sheet.max_row
            if data_start_row > data_end_row:
                raise ChinaSouthernAirBookingExcelError("南航订舱模板没有数据填写区")

            if cls.CARGO_TYPE_OPTIONS_SHEET in workbook.sheetnames:
                workbook.remove(workbook[cls.CARGO_TYPE_OPTIONS_SHEET])
            options_sheet = workbook.create_sheet(cls.CARGO_TYPE_OPTIONS_SHEET)
            for row_number, label in enumerate(labels, start=1):
                cell = options_sheet.cell(row=row_number, column=1, value=label)
                # 数据字典内容始终按普通文本写入，避免以“=”开头的异常名称
                # 被 Excel 当作公式执行。
                cell.data_type = "s"
            options_sheet.sheet_state = "veryHidden"

            if cls.CARGO_TYPE_OPTIONS_RANGE in workbook.defined_names:
                del workbook.defined_names[cls.CARGO_TYPE_OPTIONS_RANGE]
            workbook.defined_names.add(
                DefinedName(
                    cls.CARGO_TYPE_OPTIONS_RANGE,
                    attr_text=(
                        f"'{cls.CARGO_TYPE_OPTIONS_SHEET}'!$A$1:$A${len(labels)}"
                    ),
                )
            )

            validation = DataValidation(
                type="list",
                formula1=cls.CARGO_TYPE_OPTIONS_RANGE,
                allow_blank=True,
            )
            validation.errorTitle = "货物类型无效"
            validation.error = "请选择货物类型下拉框中的选项"
            validation.showErrorMessage = True
            # 不显示选中单元格时的黄色输入提示浮窗，仅保留下拉选择与
            # 非法值校验，避免提示内容遮挡模板填写区域。
            validation.showInputMessage = False
            sheet.add_data_validation(validation)
            validation.add(
                sheet.cell(row=data_start_row, column=cargo_type_column).coordinate
                + ":"
                + sheet.cell(row=data_end_row, column=cargo_type_column).coordinate
            )

            output = BytesIO()
            workbook.save(output)
            return output.getvalue()
        finally:
            workbook.close()

    @classmethod
    def _read_row(
        cls, sheet, row_number: int, header_mapping: Mapping[str, int]
    ) -> Dict[str, Any]:
        result: Dict[str, Any] = {}
        for header, field_name in cls.HEADER_FIELDS.items():
            # 出港货邮处理费选项列为可选列；列缺失与单元格留空都统一视为空值，
            # 由执行阶段决定是否透传南航查询结果。
            column_number = header_mapping.get(header)
            result[field_name] = (
                sheet.cell(row_number, column_number).value
                if column_number is not None
                else None
            )
        return result

    @classmethod
    def _is_input_row(cls, row: Mapping[str, Any]) -> bool:
        return any(cls._text(row.get(field)) for field in cls.USER_INPUT_FIELDS)

    @classmethod
    def _build_form_data(
        cls,
        row: Mapping[str, Any],
        *,
        row_number: int,
        cargo_type_codes: Mapping[str, str],
        allowed_fee_options: Iterable[str],
    ) -> Dict[str, Any]:
        values = {field: cls._text(value) for field, value in row.items()}
        values["flight_date"] = cls._date_text(row.get("flight_date"), row_number)

        missing_fields = [
            label for field, label in cls.REQUIRED_FIELDS.items() if not values.get(field)
        ]
        if missing_fields:
            raise ChinaSouthernAirBookingExcelError(
                f"第 {row_number} 行缺少必填数据：" + "、".join(missing_fields)
            )

        cargo_type_key = values["cargo_type"]
        cargo_type_code = cargo_type_codes.get(cargo_type_key)
        if not cargo_type_code:
            raise ChinaSouthernAirBookingExcelError(
                f"第 {row_number} 行货物类型“{cargo_type_key}”未在数据字典 nanfang_air_cargo_type 中配置"
            )

        selected_fee = values.get("outbound_cargo_and_mail_handling_fee_options", "")
        allowed_fees = set(allowed_fee_options)
        if selected_fee and selected_fee not in allowed_fees:
            raise ChinaSouthernAirBookingExcelError(
                f"第 {row_number} 行出港货邮处理费选项“{selected_fee}”无效，"
                "可选值为：" + "、".join(sorted(allowed_fees))
            )

        values["origin_station"] = values["origin_station"].upper()
        values["destination"] = values["destination"].upper()
        values["flight_number"] = values["flight_number"].upper()
        values["oversized_cargo"] = cls._flag_text(
            row.get("oversized_cargo"),
            row_number=row_number,
            field_label="超规货",
            yes_value="1",
            no_value="0",
            default="0",
        )
        # 现有南航映射中 no_dangerous_goods="0" 表示勾选无隐含危险品声明。
        values["no_dangerous_goods"] = cls._flag_text(
            row.get("no_dangerous_goods"),
            row_number=row_number,
            field_label="无隐含危险品",
            yes_value="0",
            no_value="1",
            default="0",
        )
        values["cargo_type_code"] = cargo_type_code
        values.pop("outbound_cargo_and_mail_handling_fee_options", None)

        return {
            "airline": "2",
            "bookings": [values],
            "outbound_cargo_and_mail_handling_fee_options": selected_fee,
        }

    @classmethod
    def parse(
        cls,
        content: bytes,
        *,
        cargo_type_codes: Mapping[str, str],
        allowed_fee_options: Iterable[str],
    ) -> List[Dict[str, Any]]:
        if not content:
            raise ChinaSouthernAirBookingExcelError("上传的Excel文件为空")
        if len(content) > cls.MAX_FILE_SIZE:
            raise ChinaSouthernAirBookingExcelError("Excel文件不能超过5MB")

        try:
            workbook = load_workbook(
                BytesIO(content), read_only=True, data_only=True, keep_links=False
            )
        except (BadZipFile, InvalidFileException, OSError, ValueError, KeyError, EOFError) as exc:
            raise ChinaSouthernAirBookingExcelError(
                "Excel文件无法解析，请使用系统下载的新版 .xlsx 模板"
            ) from exc

        try:
            sheet = workbook.worksheets[0]
            header_row, header_mapping = cls._find_header(sheet)
            if sheet.max_row > cls.MAX_WORKSHEET_ROWS:
                raise ChinaSouthernAirBookingExcelError(
                    f"Excel有效范围超过 {cls.MAX_WORKSHEET_ROWS} 行，请删除多余空白行后重试"
                )
            parsed_rows: List[Dict[str, Any]] = []
            errors: List[Dict[str, Any]] = []
            for row_number in range(header_row + 1, sheet.max_row + 1):
                raw_row = cls._read_row(sheet, row_number, header_mapping)
                if not cls._is_input_row(raw_row):
                    continue
                if len(parsed_rows) >= cls.MAX_BOOKING_ROWS:
                    raise ChinaSouthernAirBookingExcelError(
                        f"单次最多导入 {cls.MAX_BOOKING_ROWS} 条订舱数据"
                    )
                try:
                    form_data = cls._build_form_data(
                        raw_row,
                        row_number=row_number,
                        cargo_type_codes=cargo_type_codes,
                        allowed_fee_options=allowed_fee_options,
                    )
                    parsed_rows.append(
                        {"excel_row": row_number, "form_data": form_data}
                    )
                except ChinaSouthernAirBookingExcelError as exc:
                    errors.append({"excel_row": row_number, "message": str(exc)})

            if errors:
                raise ChinaSouthernAirBookingExcelError(
                    f"Excel中有 {len(errors)} 行数据校验失败，未创建任何订舱记录",
                    errors=errors,
                )
            if not parsed_rows:
                raise ChinaSouthernAirBookingExcelError("Excel中没有可导入的订舱数据")
            return parsed_rows
        finally:
            workbook.close()


china_southern_air_booking_excel_service = ChinaSouthernAirBookingExcelService()
