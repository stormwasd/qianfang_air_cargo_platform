"""
货站录单服务

用于处理深圳航空和南方航空的货站录单功能：

深圳航空：
1. 根据waybill数据填充Excel模板，条件如下：
   - 交接单（仅当 cargo_info.cargo_code == "044" 时生成）
   - 航空货物明细表（仅当 form_data.declaration_list == "0" 时生成）
   - 货物收运检查单（仅当 cargo_info.cargo_code == "044" 时生成）
   - 标签（必生成）
   - 充氧类水生动物货物收运检查单（仅当 oxygenated_aquatic_animal_goods_receipt_inspection_form_switch == "0" 时生成）
2. 将Excel转换为PDF（使用纯Python方案：openpyxl + reportlab）
3. 保存文件到指定目录
4. 更新waybill的cargo_station_record_status字段

南方航空：
1. 只有当 form_data.oxygenated_aquatic_animal_goods_receipt_inspection_form_switch 为 "0" 时才需要进行货站录单
2. 处理一个xlsx文件（充氧类水生动物货物收运检查单.xlsx）
3. 将Excel转换为PDF
4. 更新waybill的cargo_station_record_status字段
"""
import os
import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple, Dict, Any, List

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

from app.utils.airport_code_mapper import get_city_name_by_code


# 文件存储根目录（相对于项目根目录）
# 所有航司的货站录单文件统一存放在 generated_files/{waybill_id}/ 下
GENERATED_FILES_DIR = "generated_files"

# 深航Excel模板目录（相对于项目根目录）
TEMPLATE_DIR = "documents/shenzhen_air"
# 南航模板目录（相对于项目根目录）
CHINA_SOUTHERN_AIR_TEMPLATE_DIR = "documents/china_southern_air"

# ======== 深航文档类型常量 ========
DOC_TYPE_HANDOVER = "handover"  # 交接单
DOC_TYPE_CARGO_DETAIL = "cargo_detail"  # 航空货物明细表
DOC_TYPE_CARGO_CHECKLIST = "cargo_checklist"  # 货物收运检查单
DOC_TYPE_AQUATIC_ANIMAL_CHECKLIST = "aquatic_animal_checklist"  # 充氧类水生动物货物收运检查单
DOC_TYPE_LABEL = "label"  # 标签

# ======== 南航文档类型常量 ========
DOC_TYPE_CSA_AQUATIC_ANIMAL_CHECKLIST = "csa_aquatic_animal_checklist"  # 南航充氧类水生动物货物收运检查单（xlsx版）

# 深航文档类型到文件名的映射
DOC_TYPE_TO_FILENAME = {
    DOC_TYPE_HANDOVER: "交接单",
    DOC_TYPE_CARGO_DETAIL: "航空货物明细表",
    DOC_TYPE_CARGO_CHECKLIST: "货物收运检查单",
    DOC_TYPE_AQUATIC_ANIMAL_CHECKLIST: "充氧类水生动物货物收运检查单",
    DOC_TYPE_LABEL: "标签",
}

# 南航文档类型到文件名的映射
CSA_DOC_TYPE_TO_FILENAME = {
    DOC_TYPE_CSA_AQUATIC_ANIMAL_CHECKLIST: "充氧类水生动物货物收运检查单",
}


def _get_project_root() -> Path:
    """获取项目根目录"""
    # 从当前文件位置向上三级就是项目根目录
    # app/services/cargo_station_record_service.py -> app/services -> app -> 项目根目录
    return Path(__file__).parent.parent.parent


def _ensure_waybill_dir(waybill_id: int) -> Path:
    """
    确保waybill的文件存储目录存在
    
    所有航司的货站录单文件统一存放在 generated_files/{waybill_id}/ 下
    
    Args:
        waybill_id: 运单ID
    
    Returns:
        waybill的文件存储目录路径
    """
    project_root = _get_project_root()
    waybill_dir = project_root / GENERATED_FILES_DIR / str(waybill_id)
    waybill_dir.mkdir(parents=True, exist_ok=True)
    return waybill_dir


def _parse_flight_date(flight_date: str) -> Tuple[str, str, str]:
    """
    解析航班日期，返回年、月、日
    
    Args:
        flight_date: 航班日期字符串，格式如 "2025-01-15" 或 "2025/01/15"
    
    Returns:
        元组 (year, month, day)
    """
    if not flight_date:
        return "", "", ""
    
    # 尝试多种日期格式
    date_str = flight_date.strip()
    
    # 替换可能的分隔符
    date_str = date_str.replace("/", "-")
    
    parts = date_str.split("-")
    if len(parts) == 3:
        year = parts[0]
        month = parts[1].lstrip("0") or "0"  # 去除前导零，但保留至少一个字符
        day = parts[2].lstrip("0") or "0"
        return year, month, day
    
    return "", "", ""


def _replace_cell_value(ws: Worksheet, search_value: str, replace_value: str) -> int:
    """
    在工作表中查找并替换单元格值
    
    Args:
        ws: 工作表对象
        search_value: 要查找的值
        replace_value: 替换后的值
    
    Returns:
        替换的单元格数量
    """
    replaced_count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell_str = str(cell.value)
                if search_value in cell_str:
                    # 执行替换
                    cell.value = cell_str.replace(search_value, replace_value)
                    replaced_count += 1
    return replaced_count


def _convert_excel_to_pdf(excel_path: Path, pdf_path: Path) -> bool:
    """
    将Excel文件转换为PDF（纯Python实现）
    
    使用openpyxl读取Excel内容，使用reportlab生成PDF
    
    Args:
        excel_path: Excel文件路径
        pdf_path: 输出PDF文件路径
    
    Returns:
        是否转换成功
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm, cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        
        # 尝试注册中文字体
        chinese_font_name = "SimSun"
        font_registered = False
        
        # 尝试常见的中文字体路径
        font_paths = [
            "C:/Windows/Fonts/simsun.ttc",  # Windows
            "C:/Windows/Fonts/simhei.ttf",  # Windows
            "C:/Windows/Fonts/msyh.ttc",    # Windows 微软雅黑
            "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",  # Linux
            "/System/Library/Fonts/PingFang.ttc",  # macOS
        ]
        
        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    if font_path.endswith('.ttc'):
                        pdfmetrics.registerFont(TTFont(chinese_font_name, font_path, subfontIndex=0))
                    else:
                        pdfmetrics.registerFont(TTFont(chinese_font_name, font_path))
                    font_registered = True
                    break
                except Exception:
                    continue
        
        if not font_registered:
            # 如果没有找到中文字体，使用默认字体（可能无法显示中文）
            chinese_font_name = "Helvetica"
            print(f"警告：未找到中文字体，PDF中的中文可能无法正确显示")
        
        # 加载Excel文件
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # 获取Excel数据
        data = []
        merge_cells_info = []  # 存储合并单元格信息
        
        # 获取合并单元格信息
        for merged_range in ws.merged_cells.ranges:
            merge_cells_info.append({
                'min_row': merged_range.min_row,
                'max_row': merged_range.max_row,
                'min_col': merged_range.min_col,
                'max_col': merged_range.max_col
            })
        
        # 获取实际使用的行列范围
        max_row = ws.max_row
        max_col = ws.max_column
        
        # 读取数据
        for row_idx in range(1, max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # 跳过合并单元格中非左上角的单元格
                cell_value = ""
                if isinstance(cell, MergedCell):
                    # 查找这个单元格属于哪个合并区域，获取主单元格的值
                    for merge_info in merge_cells_info:
                        if (merge_info['min_row'] <= row_idx <= merge_info['max_row'] and
                            merge_info['min_col'] <= col_idx <= merge_info['max_col']):
                            if row_idx == merge_info['min_row'] and col_idx == merge_info['min_col']:
                                # 这是合并区域的左上角
                                main_cell = ws.cell(row=merge_info['min_row'], column=merge_info['min_col'])
                                cell_value = str(main_cell.value) if main_cell.value is not None else ""
                            break
                else:
                    cell_value = str(cell.value) if cell.value is not None else ""
                row_data.append(cell_value)
            data.append(row_data)
        
        wb.close()
        
        # 如果没有数据，返回失败
        if not data:
            print(f"Excel文件没有数据: {excel_path}")
            return False
        
        # 创建PDF文档（使用横向A4）
        doc = SimpleDocTemplate(
            str(pdf_path),
            pagesize=landscape(A4),
            rightMargin=10*mm,
            leftMargin=10*mm,
            topMargin=10*mm,
            bottomMargin=10*mm
        )
        
        # 创建样式
        styles = getSampleStyleSheet()
        
        # 创建中文段落样式
        chinese_style = ParagraphStyle(
            'Chinese',
            parent=styles['Normal'],
            fontName=chinese_font_name,
            fontSize=8,
            leading=10,
            wordWrap='CJK',
        )
        
        # 将数据转换为Paragraph对象以支持自动换行
        table_data = []
        for row in data:
            row_paragraphs = []
            for cell in row:
                if cell:
                    p = Paragraph(str(cell), chinese_style)
                else:
                    p = Paragraph("", chinese_style)
                row_paragraphs.append(p)
            table_data.append(row_paragraphs)
        
        # 计算列宽（根据页面宽度平均分配）
        page_width = landscape(A4)[0] - 20*mm  # 减去左右边距
        if max_col > 0:
            col_width = page_width / max_col
            col_widths = [col_width] * max_col
        else:
            col_widths = None
        
        # 创建表格
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        # 创建表格样式
        style_commands = [
            ('FONTNAME', (0, 0), (-1, -1), chinese_font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]
        
        # 处理合并单元格
        for merge_info in merge_cells_info:
            row_span = merge_info['max_row'] - merge_info['min_row'] + 1
            col_span = merge_info['max_col'] - merge_info['min_col'] + 1
            
            if row_span > 1 or col_span > 1:
                # reportlab使用0索引
                start_row = merge_info['min_row'] - 1
                start_col = merge_info['min_col'] - 1
                end_row = merge_info['max_row'] - 1
                end_col = merge_info['max_col'] - 1
                
                style_commands.append(
                    ('SPAN', (start_col, start_row), (end_col, end_row))
                )
        
        table.setStyle(TableStyle(style_commands))
        
        # 构建PDF
        elements = [table]
        doc.build(elements)
        
        return True
        
    except ImportError as e:
        print(f"缺少必要的库: {str(e)}")
        print("请安装 reportlab: pip install reportlab")
        return False
    except Exception as e:
        import traceback
        print(f"Excel转PDF失败: {str(e)}")
        print(f"错误详情: {traceback.format_exc()}")
        return False


def generate_handover_document(
    waybill_id: int,
    waybill_number: str,
    form_data: dict,
    business_config: dict
) -> Tuple[Path, Path]:
    """
    生成交接单文档
    
    生成条件：仅当 cargo_info.cargo_code == "044" 时生成
    
    Args:
        waybill_id: 运单ID
        waybill_number: 运单号
        form_data: 运单表单数据
        business_config: 业务参数配置
    
    Returns:
        元组 (Excel文件路径, PDF文件路径)
    """
    project_root = _get_project_root()
    template_path = project_root / TEMPLATE_DIR / "交接单.xlsx"
    
    waybill_dir = _ensure_waybill_dir(waybill_id)
    
    excel_filename = "交接单.xlsx"
    pdf_filename = "交接单.pdf"
    excel_path = waybill_dir / excel_filename
    pdf_path = waybill_dir / pdf_filename
    
    shutil.copy2(template_path, excel_path)
    
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # 提取数据
    flight_info = form_data.get("flight_info", {})
    cargo_info = form_data.get("cargo_info", {})
    
    flight_number = flight_info.get("flight_number", "")
    destination_code = flight_info.get("destination", "")
    destination_city = get_city_name_by_code(destination_code)
    flight_date = flight_info.get("flight_date", "")
    year, month, day = _parse_flight_date(flight_date)
    
    quantity = str(cargo_info.get("quantity", ""))
    weight = str(cargo_info.get("weight", ""))
    chargeable_weight = str(cargo_info.get("chargeable_weight", ""))
    cargo_name = str(cargo_info.get("cargo_name", ""))
    package = str(cargo_info.get("package", ""))
    
    # 从业务参数配置中获取shipper_or_agent
    shenzhen_air_config = business_config.get("shenzhen_air", {})
    document_config = shenzhen_air_config.get("document", {})
    domestic_cargo_checklist = document_config.get("domestic_cargo_checklist", {})
    shipper_or_agent = domestic_cargo_checklist.get("shipper_or_agent", "")
    
    # 使用直接单元格赋值（模板中短值如"1","2"等不适合用字符串搜索替换）
    ws['D4'] = flight_number        # ZH9505 → 航班号
    ws['F4'] = destination_city     # 上海虹桥 → 目的地城市
    ws['H4'] = year                 # 2025 → 年
    ws['I4'] = month                # 2 → 月
    ws['J4'] = day                  # 18 → 日
    ws['A12'] = waybill_number      # 479-53957562 → 运单号
    ws['C12'] = quantity            # 1 → 件数
    ws['D12'] = weight              # 5 → 毛重
    ws['E12'] = chargeable_weight   # 5 → 计费重量
    ws['J12'] = package             # 纸箱 → 包装
    ws['F13'] = cargo_name          # NK细胞 → 货物品名
    ws['B47'] = shipper_or_agent    # 唐文旭 → 托运人代理人
    
    # airline_consent_certificate 条件替换：非空时替换 H30 单元格
    airline_consent_certificate = form_data.get("airline_consent_certificate", "")
    if airline_consent_certificate and len(airline_consent_certificate.strip()) > 0:
        ws['H30'] = airline_consent_certificate
    
    wb.save(excel_path)
    wb.close()
    
    _convert_excel_to_pdf(excel_path, pdf_path)
    
    return excel_path, pdf_path


def generate_cargo_detail_document(
    waybill_id: int,
    waybill_number: str,
    form_data: dict,
    business_config: dict
) -> Tuple[Path, Path]:
    """
    生成航空货物明细表文档
    
    生成条件：仅当 form_data.declaration_list == "0" 时生成
    
    Args:
        waybill_id: 运单ID
        waybill_number: 运单号
        form_data: 运单表单数据
        business_config: 业务参数配置
    
    Returns:
        元组 (Excel文件路径, PDF文件路径)
    """
    project_root = _get_project_root()
    template_path = project_root / TEMPLATE_DIR / "航空货物明细表.xlsx"
    
    waybill_dir = _ensure_waybill_dir(waybill_id)
    
    excel_filename = "航空货物明细表.xlsx"
    pdf_filename = "航空货物明细表.pdf"
    excel_path = waybill_dir / excel_filename
    pdf_path = waybill_dir / pdf_filename
    
    shutil.copy2(template_path, excel_path)
    
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # 提取数据
    flight_info = form_data.get("flight_info", {})
    cargo_info = form_data.get("cargo_info", {})
    
    flight_number = flight_info.get("flight_number", "")
    flight_date = flight_info.get("flight_date", "")
    quantity = str(cargo_info.get("quantity", ""))
    weight = str(cargo_info.get("weight", ""))
    
    # 使用直接单元格赋值（避免短值"18"/"296"的误匹配风险）
    ws['B3'] = flight_number              # ZH9949 → 航班号
    ws['D3'] = waybill_number             # 479-57110642 → 主单号
    ws['G3'] = f"{quantity}件{weight}KG"  # 18件296KG → 件数重量
    
    # H3 是日期（模板中为Excel日期序列号45909），设置为实际航班日期
    try:
        date_obj = datetime.strptime(flight_date, "%Y-%m-%d")
        ws['H3'] = date_obj
    except (ValueError, TypeError):
        ws['H3'] = flight_date
    
    wb.save(excel_path)
    wb.close()
    
    _convert_excel_to_pdf(excel_path, pdf_path)
    
    return excel_path, pdf_path


def generate_cargo_checklist_document(
    waybill_id: int,
    waybill_number: str,
    form_data: dict,
    business_config: dict
) -> Tuple[Path, Path]:
    """
    生成货物收运检查单文档
    
    生成条件：仅当 cargo_info.cargo_code == "044" 时生成
    
    Args:
        waybill_id: 运单ID
        waybill_number: 运单号
        form_data: 运单表单数据
        business_config: 业务参数配置
    
    Returns:
        元组 (Excel文件路径, PDF文件路径)
    """
    project_root = _get_project_root()
    template_path = project_root / TEMPLATE_DIR / "货物收运检查单.xlsx"
    
    waybill_dir = _ensure_waybill_dir(waybill_id)
    
    excel_filename = "货物收运检查单.xlsx"
    pdf_filename = "货物收运检查单.pdf"
    excel_path = waybill_dir / excel_filename
    pdf_path = waybill_dir / pdf_filename
    
    shutil.copy2(template_path, excel_path)
    
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # 提取数据
    flight_info = form_data.get("flight_info", {})
    cargo_info = form_data.get("cargo_info", {})
    
    flight_number = flight_info.get("flight_number", "")
    destination_code = flight_info.get("destination", "")
    destination_city = get_city_name_by_code(destination_code)
    quantity = str(cargo_info.get("quantity", ""))
    weight = str(cargo_info.get("weight", ""))
    chargeable_weight = str(cargo_info.get("chargeable_weight", ""))
    cargo_name = str(cargo_info.get("cargo_name", ""))
    package = str(cargo_info.get("package", ""))
    
    # 使用直接单元格赋值（避免短值如"110","400"的误匹配风险）
    ws['E4'] = flight_number          # ZH9929 → 航班号
    ws['M4'] = destination_city       # 济南 → 目的地
    ws['A8'] = waybill_number         # 479-57515651 → 运单号
    ws['E8'] = quantity               # 110 → 件数
    ws['G8'] = weight                 # 400 → 毛重
    ws['I8'] = chargeable_weight      # 400 → 计费重量
    ws['K8'] = cargo_name             # 货物品名
    ws['P8'] = package                # 纸箱 → 包装
    
    wb.save(excel_path)
    wb.close()
    
    _convert_excel_to_pdf(excel_path, pdf_path)
    
    return excel_path, pdf_path


def _fill_aquatic_animal_checklist_xlsx(ws: Worksheet, waybill_number: str,
                                        form_data: dict, shipper_unit: str):
    """
    填充充氧类水生动物货物收运检查单Excel工作表（深航和南航共用同一模板结构）
    
    使用直接单元格赋值确保精确替换，避免短值误匹配。
    """
    flight_info = form_data.get("flight_info", {})
    cargo_info = form_data.get("cargo_info", {})
    
    flight_number = flight_info.get("flight_number", "")
    destination_code = flight_info.get("destination", "")
    destination_city = get_city_name_by_code(destination_code)
    flight_date = flight_info.get("flight_date", "")
    quantity = str(cargo_info.get("quantity", ""))
    weight = str(cargo_info.get("weight", ""))
    oxygen_supply_test_results = form_data.get("oxygen_supply_test_results", "")
    
    # 基本信息（直接单元格赋值）
    ws['D3'] = shipper_unit                  # 深圳丰德航空物流有限公司 → 托运代理人
    ws['C4'] = waybill_number                # 479-60491104 → 运单号
    ws['I4'] = f"{quantity}件{weight}KG"     # 3件80KG → 件数/重量
    ws['C5'] = flight_number                 # CA4336 → 航班号
    ws['I5'] = destination_city              # 成都双流 → 目的站
    
    # I3 是日期（模板中为Excel日期序列号46086），设置为实际航班日期
    try:
        date_obj = datetime.strptime(flight_date, "%Y-%m-%d")
        ws['I3'] = date_obj
    except (ValueError, TypeError):
        ws['I3'] = flight_date
    
    # 新增替换项
    ws['C16'] = quantity                     # 3 → 开包件数
    ws['F16'] = oxygen_supply_test_results   # 蔬菜、丝瓜、芥菜 → 检查结果
    ws['C18'] = flight_date                  # 填航班日期 → 航班日期


def generate_aquatic_animal_checklist_document(
    waybill_id: int,
    waybill_number: str,
    form_data: dict,
    business_config: dict
) -> Tuple[Path, Path]:
    """
    生成充氧类水生动物货物收运检查单文档（深航版）
    
    生成条件：仅当 oxygenated_aquatic_animal_goods_receipt_inspection_form_switch == "0" 时生成
    
    Args:
        waybill_id: 运单ID
        waybill_number: 运单号
        form_data: 运单表单数据
        business_config: 业务参数配置
    
    Returns:
        元组 (Excel文件路径, PDF文件路径)
    """
    project_root = _get_project_root()
    template_path = project_root / TEMPLATE_DIR / "充氧类水生动物货物收运检查单.xlsx"
    
    waybill_dir = _ensure_waybill_dir(waybill_id)
    
    excel_filename = "充氧类水生动物货物收运检查单.xlsx"
    pdf_filename = "充氧类水生动物货物收运检查单.pdf"
    excel_path = waybill_dir / excel_filename
    pdf_path = waybill_dir / pdf_filename
    
    shutil.copy2(template_path, excel_path)
    
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # 深航托运单位来源于 shipper_consignee_info
    shipper_consignee_info = form_data.get("shipper_consignee_info", {})
    shipper_unit = shipper_consignee_info.get("shipper_unit", "")
    
    _fill_aquatic_animal_checklist_xlsx(ws, waybill_number, form_data, shipper_unit)
    
    wb.save(excel_path)
    wb.close()
    
    _convert_excel_to_pdf(excel_path, pdf_path)
    
    return excel_path, pdf_path


def generate_label_document(
    waybill_id: int,
    waybill_number: str,
    form_data: dict,
    business_config: dict
) -> Tuple[Path, Path]:
    """
    生成标签文档（必生成）
    
    模板中第一个标签区域的数据为源数据，后续标签通过公式引用自动同步。
    只需修改第一个标签区域（前5行数据行）即可。
    
    Args:
        waybill_id: 运单ID
        waybill_number: 运单号
        form_data: 运单表单数据
        business_config: 业务参数配置
    
    Returns:
        元组 (Excel文件路径, PDF文件路径)
    """
    project_root = _get_project_root()
    template_path = project_root / TEMPLATE_DIR / "标签.xlsx"
    
    waybill_dir = _ensure_waybill_dir(waybill_id)
    
    excel_filename = "标签.xlsx"
    pdf_filename = "标签.pdf"
    excel_path = waybill_dir / excel_filename
    pdf_path = waybill_dir / pdf_filename
    
    shutil.copy2(template_path, excel_path)
    
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # 提取数据
    flight_info = form_data.get("flight_info", {})
    cargo_info = form_data.get("cargo_info", {})
    
    flight_number = flight_info.get("flight_number", "")
    origin_station_code = flight_info.get("origin_station", "")
    destination_code = flight_info.get("destination", "")
    origin_city = get_city_name_by_code(origin_station_code)
    destination_city = get_city_name_by_code(destination_code)
    quantity = str(cargo_info.get("quantity", ""))
    weight = str(cargo_info.get("weight", ""))
    
    # 只需修改第一个标签区域，后续标签通过公式自动同步
    ws['C2'] = waybill_number    # 479-58183392 → 航班运单号
    ws['B3'] = quantity          # 6 → 件数
    ws['D3'] = weight            # 36 → 重量
    ws['A4'] = origin_city       # 深圳 → 始发站城市
    ws['C4'] = destination_city  # 合肥 → 目的站城市
    ws['D4'] = flight_number     # ZH9945 → 航班号
    
    wb.save(excel_path)
    wb.close()
    
    _convert_excel_to_pdf(excel_path, pdf_path)
    
    return excel_path, pdf_path


def generate_all_documents(
    waybill_id: int,
    waybill_number: str,
    form_data: dict,
    business_config: dict
) -> Dict[str, Dict[str, Optional[str]]]:
    """
    生成深航所有货站录单文档
    
    根据 form_data 中的配置决定生成哪些文档：
    - 交接单（仅当 cargo_info.cargo_code == "044" 时生成）
    - 航空货物明细表（仅当 form_data.declaration_list == "0" 时生成）
    - 货物收运检查单（仅当 cargo_info.cargo_code == "044" 时生成）
    - 标签（必生成）
    - 充氧类水生动物货物收运检查单（仅当 oxygenated_aquatic_animal_goods_receipt_inspection_form_switch == "0" 时生成）
    
    Args:
        waybill_id: 运单ID
        waybill_number: 运单号
        form_data: 运单表单数据
        business_config: 业务参数配置
    
    Returns:
        生成结果字典
    """
    results = {}
    cargo_info = form_data.get("cargo_info", {})
    cargo_code = str(cargo_info.get("cargo_code", ""))
    declaration_list = str(form_data.get("declaration_list", ""))
    aquatic_switch = form_data.get("oxygenated_aquatic_animal_goods_receipt_inspection_form_switch", "1")
    
    # 生成交接单（仅当 cargo_code == "044"）
    if cargo_code == "044":
        try:
            excel_path, pdf_path = generate_handover_document(
                waybill_id, waybill_number, form_data, business_config
            )
            results[DOC_TYPE_HANDOVER] = {
                "excel": str(excel_path),
                "pdf": str(pdf_path) if pdf_path.exists() else None
            }
        except Exception as e:
            print(f"生成交接单失败: {str(e)}")
            results[DOC_TYPE_HANDOVER] = {"excel": None, "pdf": None, "error": str(e)}
    
    # 生成航空货物明细表（仅当 declaration_list == "0"）
    if declaration_list == "0":
        try:
            excel_path, pdf_path = generate_cargo_detail_document(
                waybill_id, waybill_number, form_data, business_config
            )
            results[DOC_TYPE_CARGO_DETAIL] = {
                "excel": str(excel_path),
                "pdf": str(pdf_path) if pdf_path.exists() else None
            }
        except Exception as e:
            print(f"生成航空货物明细表失败: {str(e)}")
            results[DOC_TYPE_CARGO_DETAIL] = {"excel": None, "pdf": None, "error": str(e)}
    
    # 生成货物收运检查单（仅当 cargo_code == "044"）
    if cargo_code == "044":
        try:
            excel_path, pdf_path = generate_cargo_checklist_document(
                waybill_id, waybill_number, form_data, business_config
            )
            results[DOC_TYPE_CARGO_CHECKLIST] = {
                "excel": str(excel_path),
                "pdf": str(pdf_path) if pdf_path.exists() else None
            }
        except Exception as e:
            print(f"生成货物收运检查单失败: {str(e)}")
            results[DOC_TYPE_CARGO_CHECKLIST] = {"excel": None, "pdf": None, "error": str(e)}
    
    # 生成标签（必生成）
    try:
        excel_path, pdf_path = generate_label_document(
            waybill_id, waybill_number, form_data, business_config
        )
        results[DOC_TYPE_LABEL] = {
            "excel": str(excel_path),
            "pdf": str(pdf_path) if pdf_path.exists() else None
        }
    except Exception as e:
        print(f"生成标签失败: {str(e)}")
        results[DOC_TYPE_LABEL] = {"excel": None, "pdf": None, "error": str(e)}
    
    # 生成充氧类水生动物货物收运检查单（仅当开关为"0"时生成）
    if aquatic_switch == "0":
        try:
            excel_path, pdf_path = generate_aquatic_animal_checklist_document(
                waybill_id, waybill_number, form_data, business_config
            )
            results[DOC_TYPE_AQUATIC_ANIMAL_CHECKLIST] = {
                "excel": str(excel_path),
                "pdf": str(pdf_path) if pdf_path.exists() else None
            }
        except Exception as e:
            print(f"生成充氧类水生动物货物收运检查单失败: {str(e)}")
            results[DOC_TYPE_AQUATIC_ANIMAL_CHECKLIST] = {"excel": None, "pdf": None, "error": str(e)}
    
    return results


def get_document_path(
    waybill_id: int,
    doc_type: str,
    file_format: str = "pdf"
) -> Optional[Path]:
    """
    获取指定运单的指定文档路径
    
    Args:
        waybill_id: 运单ID
        doc_type: 文档类型 (handover, cargo_detail, cargo_checklist, label, aquatic_animal_checklist)
        file_format: 文件格式 (pdf 或 excel)
    
    Returns:
        文件路径，如果不存在则返回None
    """
    project_root = _get_project_root()
    waybill_dir = project_root / GENERATED_FILES_DIR / str(waybill_id)
    
    if not waybill_dir.exists():
        return None
    
    # 获取文档名称
    doc_name = DOC_TYPE_TO_FILENAME.get(doc_type)
    if not doc_name:
        return None
    
    # 确定文件扩展名
    extension = ".pdf" if file_format == "pdf" else ".xlsx"
    
    # 使用固定文件名（不带时间戳）
    file_path = waybill_dir / f"{doc_name}{extension}"
    
    if file_path.exists():
        return file_path
    
    return None


def list_documents(waybill_id: int) -> Dict[str, Dict[str, Optional[str]]]:
    """
    列出指定运单的所有已生成文档（深航）
    
    Args:
        waybill_id: 运单ID
    
    Returns:
        文档列表字典，格式如：
        {
            "handover": {"excel": "/path/to/excel", "pdf": "/path/to/pdf"},
            "cargo_detail": {"excel": "/path/to/excel", "pdf": "/path/to/pdf"},
            "cargo_checklist": {"excel": "/path/to/excel", "pdf": "/path/to/pdf"},
            "label": {"excel": "/path/to/excel", "pdf": "/path/to/pdf"},
            "aquatic_animal_checklist": {"excel": "/path/to/excel", "pdf": "/path/to/pdf"}
        }
    """
    results = {}
    
    # 遍历所有文档类型
    for doc_type in [DOC_TYPE_HANDOVER, DOC_TYPE_CARGO_DETAIL, DOC_TYPE_CARGO_CHECKLIST, DOC_TYPE_LABEL, DOC_TYPE_AQUATIC_ANIMAL_CHECKLIST]:
        excel_path = get_document_path(waybill_id, doc_type, "excel")
        pdf_path = get_document_path(waybill_id, doc_type, "pdf")
        
        # 只有当文件存在时才添加到结果中
        if excel_path or pdf_path:
            results[doc_type] = {
                "excel": str(excel_path) if excel_path else None,
                "pdf": str(pdf_path) if pdf_path else None
            }
    
    return results


# ======== 南航货站录单相关函数 ========
# 注意：南航和深航的文件统一存放在 generated_files/{waybill_id}/ 下


def generate_csa_aquatic_animal_checklist_document(
    waybill_id: int,
    waybill_number: str,
    form_data: dict,
    business_config: dict
) -> Tuple[Path, Path]:
    """
    生成南航充氧类水生动物货物收运检查单文档（xlsx格式）
    
    生成条件：仅当 oxygenated_aquatic_animal_goods_receipt_inspection_form_switch == "0" 时生成
    使用与深航相同结构的xlsx模板。
    
    Args:
        waybill_id: 运单ID
        waybill_number: 运单号
        form_data: 运单表单数据
        business_config: 业务参数配置
    
    Returns:
        元组 (Excel文件路径, PDF文件路径)
    """
    project_root = _get_project_root()
    template_path = project_root / CHINA_SOUTHERN_AIR_TEMPLATE_DIR / "充氧类水生动物货物收运检查单.xlsx"
    
    if not template_path.exists():
        raise FileNotFoundError(f"南航模板文件不存在: {template_path}")
    
    waybill_dir = _ensure_waybill_dir(waybill_id)
    
    excel_filename = "充氧类水生动物货物收运检查单.xlsx"
    pdf_filename = "充氧类水生动物货物收运检查单.pdf"
    excel_path = waybill_dir / excel_filename
    pdf_path = waybill_dir / pdf_filename
    
    shutil.copy2(template_path, excel_path)
    
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # 南航托运单位来源于 contact_info
    contact_info = form_data.get("contact_info", {})
    shipper_unit = contact_info.get("shipper_unit", "")
    
    _fill_aquatic_animal_checklist_xlsx(ws, waybill_number, form_data, shipper_unit)
    
    wb.save(excel_path)
    wb.close()
    
    _convert_excel_to_pdf(excel_path, pdf_path)
    
    return excel_path, pdf_path


def generate_csa_all_documents(
    waybill_id: int,
    waybill_number: str,
    form_data: dict,
    business_config: dict
) -> Dict[str, Dict[str, Optional[str]]]:
    """
    生成南航所有货站录单文档
    
    南航货站录单只有当 oxygenated_aquatic_animal_goods_receipt_inspection_form_switch 为 "0" 时才需要生成
    生成文档：充氧类水生动物货物收运检查单.xlsx（及对应PDF）
    
    Args:
        waybill_id: 运单ID
        waybill_number: 运单号
        form_data: 运单表单数据
        business_config: 业务参数配置
    
    Returns:
        生成结果字典
    """
    results = {}
    
    aquatic_switch = form_data.get("oxygenated_aquatic_animal_goods_receipt_inspection_form_switch", "1")
    if aquatic_switch != "0":
        return results
    
    try:
        excel_path, pdf_path = generate_csa_aquatic_animal_checklist_document(
            waybill_id, waybill_number, form_data, business_config
        )
        results[DOC_TYPE_CSA_AQUATIC_ANIMAL_CHECKLIST] = {
            "excel": str(excel_path),
            "pdf": str(pdf_path) if pdf_path.exists() else None
        }
    except Exception as e:
        print(f"生成南航充氧类水生动物货物收运检查单失败: {str(e)}")
        results[DOC_TYPE_CSA_AQUATIC_ANIMAL_CHECKLIST] = {"excel": None, "pdf": None, "error": str(e)}
    
    return results


def get_csa_document_path(
    waybill_id: int,
    doc_type: str,
    file_format: str = "xlsx"
) -> Optional[Path]:
    """
    获取指定南航运单的指定文档路径
    
    Args:
        waybill_id: 运单ID
        doc_type: 文档类型 (csa_aquatic_animal_checklist)
        file_format: 文件格式 (xlsx 或 pdf)
    
    Returns:
        文件路径，如果不存在则返回None
    """
    project_root = _get_project_root()
    waybill_dir = project_root / GENERATED_FILES_DIR / str(waybill_id)
    
    if not waybill_dir.exists():
        return None
    
    doc_name = CSA_DOC_TYPE_TO_FILENAME.get(doc_type)
    if not doc_name:
        return None
    
    extension = ".pdf" if file_format == "pdf" else ".xlsx"
    file_path = waybill_dir / f"{doc_name}{extension}"
    
    if file_path.exists():
        return file_path
    
    return None


def list_csa_documents(waybill_id: int) -> Dict[str, Dict[str, Optional[str]]]:
    """
    列出指定南航运单的所有已生成文档
    
    Args:
        waybill_id: 运单ID
    
    Returns:
        文档列表字典
    """
    results = {}
    
    for doc_type in [DOC_TYPE_CSA_AQUATIC_ANIMAL_CHECKLIST]:
        excel_path = get_csa_document_path(waybill_id, doc_type, "xlsx")
        pdf_path = get_csa_document_path(waybill_id, doc_type, "pdf")
        
        if excel_path or pdf_path:
            results[doc_type] = {
                "excel": str(excel_path) if excel_path else None,
                "pdf": str(pdf_path) if pdf_path else None
            }
    
    return results


def is_csa_cargo_station_record_required(form_data: dict) -> bool:
    """
    判断南航是否需要进行货站录单
    
    只有当 oxygenated_aquatic_animal_goods_receipt_inspection_form_switch 为 "0" 时才需要
    
    Args:
        form_data: 运单表单数据
    
    Returns:
        是否需要进行货站录单
    """
    aquatic_switch = form_data.get("oxygenated_aquatic_animal_goods_receipt_inspection_form_switch", "1")
    return aquatic_switch == "0"
