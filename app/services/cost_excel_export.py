"""费用登记 Excel 导出的表头及字段展示转换。"""
from typing import Dict, List, Tuple

from openpyxl.worksheet.worksheet import Worksheet


BILL_OF_LADING_EXPORT_LABELS: Dict[str, str] = {
    "1-0": "一主",
    "1-1": "一主（一）分",
    "1-2": "一主（二）分",
    "1-3": "一主（三）分",
    "1-4": "一主（四）分",
    "1-5": "一主（五）分",
    "1-6": "一主（六）分",
    "1-7": "一主（七）分",
    "1-8": "一主（八）分",
    "1-9": "一主（九）分",
    "2-0": "直单（虚拟分单）",
    "2-1": "直单（虚拟分单*1）",
    "2-2": "直单（虚拟分单*2）",
    "2-3": "直单（虚拟分单*3）",
    "2-4": "直单（虚拟分单*4）",
    "2-5": "直单（虚拟分单*5）",
    "2-6": "直单（虚拟分单*6）",
    "2-7": "直单（虚拟分单*7）",
    "2-8": "直单（虚拟分单*8）",
    "2-9": "直单（虚拟分单*9）",
}

BILL_OF_LADING_STORED_PREFIXES: Dict[str, str] = {
    "一主多分": "1",
    "直单": "2",
}


def format_bill_of_lading_for_export(value: object) -> str:
    """将前端提单编码转换为页面展示名称，未知值保持原样。"""
    if value is None:
        return ""

    raw_value = str(value)
    lookup_key = raw_value.strip()
    label = BILL_OF_LADING_EXPORT_LABELS.get(lookup_key)
    if label is not None:
        return label

    # 新增、修改接口会原样保存 bill_of_lading。当前前端也可能提交
    # “一主多分-6”或“直单-3”，需与“1-6”或“2-3”采用同一展示规则。
    # 仅转换已知前缀，避免误改真实运单号或其他未知业务值。
    prefix, separator, sequence = lookup_key.rpartition("-")
    code_prefix = BILL_OF_LADING_STORED_PREFIXES.get(prefix.strip())
    sequence = sequence.strip()
    if separator and code_prefix and sequence.isdigit():
        normalized_key = f"{code_prefix}-{int(sequence)}"
        label = BILL_OF_LADING_EXPORT_LABELS.get(normalized_key)
        if label is not None:
            return label

    return raw_value


COST_EXPORT_HEADERS: Tuple[str, ...] = (
    # 货主托运信息（1-17）
    "制单时间", "内部单据ID", "进仓日期", "客户名称", "始发站-目的站",
    "报关", "提单", "航班日期", "航班号", "航班单号",
    "件数", "实际重量(kg)", "计费重量(kg)", "体积(m³)", "一程重量(kg)",
    "代理", "委托备注",

    # 应收款项（18-36）
    "应收-单价", "应收-运费计算方式", "应收-运费", "应收-提单费/信息录入费", "应收-分单费 电报费/底账费",
    "应收-报关费", "应收-续页费", "应收-海关查验费", "应收-磁检费/安检费",
    "应收-TC费", "应收-前置仓费", "应收-制单费",
    "应收-制单分单费", "应收-垫板费", "应收-打板/装箱费", "应收-探板费",
    "应收-耗材费", "应收-一程费用", "应收-合计",

    # 应付款项 - 国际空运（37-59）
    "国空应付-小计", "国空应付-外发单位", "国空应付-始发站",
    "国空应付-到达站", "国空应付-航班单号", "国空应付-航班号",
    "国空应付-航班日期", "国空应付-件数", "国空应付-实际重量", "国空应付-体积",
    "国空应付-计费重量", "国空应付-单价", "国空应付-运费", "国空应付-提单费",
    "国空应付-分单费", "国空应付-燃油费", "国空应付-TC费",
    "国空应付-报关费", "国空应付-续页费", "国空应付-耗材费", "国空应付-前置仓",
    "国空应付-其他费用", "国空应付-备注",

    # 应付款项 - 汽运（60-70）
    "汽运应付-小计", "汽运应付-托运日期", "汽运应付-外发单位", "汽运应付-件数",
    "汽运应付-重量", "汽运应付-体积", "汽运应付-单价", "汽运应付-运费",
    "汽运应付-制单费", "汽运应付-其他费用", "汽运应付-备注",

    # 应付款项 - 国内空运（71-87）
    "国空内应付-小计", "国空内应付-托运日期", "国空内应付-外发单位", "国空内应付-始发站",
    "国空内应付-到达站", "国空内应付-航空公司", "国空内应付-航空单位", "国空内应付-航空单号",
    "国空内应付-航班号", "国空内应付-航班日期", "国空内应付-件数", "国空内应付-实际重量",
    "国空内应付-计费重量", "国空内应付-费率", "国空内应付-运费", "国空内应付-其他费用",
    "国空内应付-备注",

    # 应付款项 - 报关（88-95）
    "报关应付-小计", "报关应付-报关日期", "报关应付-报关代理", "报关应付-报关费",
    "报关应付-续页费", "报关应付-查验/删单费", "报关应付-其他费用", "报关应付-备注",

    # 应付款项 - 地面操作（96-106）
    "地面应付-小计", "地面应付-托运日期", "地面应付-外发单位", "地面应付-计费重量",
    "地面应付-费率", "地面应付-运费", "地面应付-提单/快件处置费", "地面应付-安检/报关费",
    "地面应付-打板/退场费", "地面应付-其他费用", "地面应付-备注",

    # 应付款项总计、折让信息、业务信息、经营信息（107-113）
    "应付合计", "折让人员", "折让费", "业务员", "提成金额", "利润", "利润率(%)",
)


_LEAF_PREFIXES: Tuple[str, ...] = (
    "应收-",
    "国空应付-",
    "汽运应付-",
    "国空内应付-",
    "报关应付-",
    "地面应付-",
)


def _leaf_header(raw_header: str) -> str:
    if raw_header == "委托备注":
        return "备注"
    if raw_header == "应付合计":
        # 第 106 列没有下一层字段，标题在第二、三行纵向合并展示。
        return ""
    for prefix in _LEAF_PREFIXES:
        if raw_header.startswith(prefix):
            return raw_header[len(prefix):]
    return raw_header


def append_cost_export_headers(ws: Worksheet) -> List[str]:
    """写入截图所示的三级分组表头，并返回最底层字段标题。"""
    if len(COST_EXPORT_HEADERS) != 113:
        raise RuntimeError("费用登记导出字段数量异常，预期为 113 列")

    leaf_headers = [_leaf_header(header) for header in COST_EXPORT_HEADERS]
    top_headers: List[str] = [""] * len(COST_EXPORT_HEADERS)
    subgroup_headers: List[str] = [""] * len(COST_EXPORT_HEADERS)

    # 一级分组：货主托运、应收、应付、折让、业务、经营。
    for index, title in (
        (0, "货主托运信息"),
        (17, "应收款项"),
        (36, "应付款项"),
        (107, "折让信息"),
        (109, "业务信息"),
        (111, "经营信息"),
    ):
        top_headers[index] = title

    # 应付款项下的二级分组。
    for index, title in (
        (36, "国际空运信息"),
        (59, "汽运信息"),
        (70, "国内空运信息"),
        (87, "报关信息"),
        (95, "地面操作信息"),
        (106, "应付合计"),
    ):
        subgroup_headers[index] = title

    ws.append(top_headers)
    ws.append(subgroup_headers)
    ws.append(leaf_headers)

    # 无二级分组的一级标题跨前两行；应付款项保留完整三级结构。
    for start_col, end_col in (
        (1, 17),
        (18, 36),
        (108, 109),
        (110, 111),
        (112, 113),
    ):
        ws.merge_cells(
            start_row=1,
            start_column=start_col,
            end_row=2,
            end_column=end_col,
        )
    ws.merge_cells(start_row=1, start_column=37, end_row=1, end_column=107)

    for start_col, end_col in ((37, 59), (60, 70), (71, 87), (88, 95), (96, 106)):
        ws.merge_cells(
            start_row=2,
            start_column=start_col,
            end_row=2,
            end_column=end_col,
        )
    ws.merge_cells(start_row=2, start_column=107, end_row=3, end_column=107)

    return leaf_headers
