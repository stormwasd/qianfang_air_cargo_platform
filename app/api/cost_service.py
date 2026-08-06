"""
费用登记台 API 接口（层级化结构支持）
"""
import io
from datetime import datetime, date
from typing import List, Optional, Any, Dict

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from fastapi import APIRouter, Depends, Path, Query
from fastapi.responses import Response
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.database import get_db
from app.core.exceptions import NotFoundException, BadRequestException
from app.core.response import success_response
from app.api.deps import get_current_active_user
from app.models.user import User
from app.models.cost_service import CostRegistration, CostConsignment
from app.schemas.cost_service import (
    CostRegistrationSave,
    CostConsignmentCreate,
    CostConsignmentUpdate,
    CostConsignmentQuery,
    CostBatchDeleteRequest,
    CostExportExcelRequest,
)
from app.utils.helpers import format_datetime_china, get_china_now

router = APIRouter()


def _parse_datetime(val: Any) -> Optional[datetime]:
    """解析日期时间"""
    if not val:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        val_str = val.strip()
        if not val_str:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(val_str, fmt)
            except ValueError:
                continue
    return None


def _parse_date(val: Any) -> Optional[date]:
    """解析日期"""
    if not val:
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, str):
        val_str = val.strip()
        if not val_str:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(val_str, fmt).date()
            except ValueError:
                continue
    return None


def _format_cost_record(record: Any) -> Dict[str, Any]:
    """将 SQLAlchemy 记录格式化为包含层级关系的字典响应格式"""
    if not record:
        return {}
    
    def _to_float(v):
        return float(v) if v is not None else None

    def _to_date_str(v):
        if not v:
            return None
        if isinstance(v, (datetime, date)):
            return v.strftime("%Y-%m-%d")
        return str(v)

    def _to_dt_str(v):
        if not v:
            return None
        if isinstance(v, datetime):
            return v.strftime("%Y-%m-%d %H:%M:%S")
        return str(v)

    data = {
        "id": str(record.id),
        
        # (1) 货主委托信息
        "consignor_info": {
            "create_time": _to_dt_str(record.create_time),
            "internal_doc_id": record.internal_doc_id or "",
            "warehouse_entry_date": _to_date_str(record.warehouse_entry_date),
            "customer_name": record.customer_name or "",
            "origin_destination": record.origin_destination or "",
            "customs_declaration": record.customs_declaration or "",
            "bill_of_lading": record.bill_of_lading or "",
            "flight_date": _to_date_str(record.flight_date),
            "flight_no": record.flight_no or "",
            "flight_doc_no": record.flight_doc_no or "",
            "pieces": record.pieces,
            "actual_weight": _to_float(record.actual_weight),
            "chargeable_weight": _to_float(record.chargeable_weight),
            "volume": _to_float(record.volume),
            "first_leg_weight": _to_float(record.first_leg_weight),
            "agent": record.agent or "",
            "remark": record.remark or "",
        },
        
        # (2) 应收款项
        "receivables": {
            "unit_price": _to_float(record.unit_price),
            "freight": _to_float(record.receivable_freight),
            "lading_info_fee": _to_float(record.receivable_lading_info_fee),
            "split_offset_telex_fee": _to_float(record.receivable_split_offset_telex_fee),
            "customs_fee": _to_float(record.receivable_customs_fee),
            "continuation_sheet_fee": _to_float(record.receivable_continuation_sheet_fee),
            "customs_inspection_fee": _to_float(record.receivable_customs_inspection_fee),
            "magnetic_security_fee": _to_float(record.receivable_magnetic_security_fee),
            "tc_express_fee": _to_float(record.receivable_tc_express_fee),
            "warehouse_ground_fee": _to_float(record.receivable_warehouse_ground_fee),
            "doc_make_fee": _to_float(record.receivable_doc_make_fee),
            "doc_split_fee": _to_float(record.receivable_doc_split_fee),
            "skid_fee": _to_float(record.receivable_skid_fee),
            "pallet_packing_fee": _to_float(record.receivable_pallet_packing_fee),
            "probe_fee": _to_float(record.receivable_probe_fee),
            "consumables_fee": _to_float(record.receivable_consumables_fee),
            "first_leg_fee": _to_float(record.receivable_first_leg_fee),
            "total": _to_float(record.receivable_total),
        },
        
        # (3) 应付款项
        "payables": {
            "intl_air": {
                "subtotal": _to_float(record.pay_intl_air_subtotal),
                "date": _to_date_str(record.pay_intl_air_date),
                "outsource_unit": record.pay_intl_air_outsource_unit or "",
                "origin": record.pay_intl_air_origin or "",
                "destination": record.pay_intl_air_destination or "",
                "airline": record.pay_intl_air_airline or "",
                "flight_doc_no": record.pay_intl_air_flight_doc_no or "",
                "flight_no": record.pay_intl_air_flight_no or "",
                "flight_date": _to_date_str(record.pay_intl_air_flight_date),
                "pieces": record.pay_intl_air_pieces,
                "weight": _to_float(record.pay_intl_air_weight),
                "volume": _to_float(record.pay_intl_air_volume),
                "chargeable_weight": _to_float(record.pay_intl_air_chargeable_weight),
                "rate": _to_float(record.pay_intl_air_rate),
                "freight": _to_float(record.pay_intl_air_freight),
                "lading_fee": _to_float(record.pay_intl_air_lading_fee),
                "split_fee": _to_float(record.pay_intl_air_split_fee),
                "borrow_magnetic_fuel_pickup_fee": _to_float(record.pay_intl_air_borrow_magnetic_fuel_pickup_fee),
                "tc_network_disposal_fee": _to_float(record.pay_intl_air_tc_network_disposal_fee),
                "customs_fee": _to_float(record.pay_intl_air_customs_fee),
                "continuation_sheet_fee": _to_float(record.pay_intl_air_continuation_sheet_fee),
                "consumables_fee": _to_float(record.pay_intl_air_consumables_fee),
                "front_warehouse": _to_float(record.pay_intl_air_front_warehouse),
                "other_fee": _to_float(record.pay_intl_air_other_fee),
                "remark": record.pay_intl_air_remark or "",
            },
            "trucking": {
                "subtotal": _to_float(record.pay_trucking_subtotal),
                "date": _to_date_str(record.pay_trucking_date),
                "outsource_unit": record.pay_trucking_outsource_unit or "",
                "pieces": record.pay_trucking_pieces,
                "weight": _to_float(record.pay_trucking_weight),
                "volume": _to_float(record.pay_trucking_volume),
                "unit_price": _to_float(record.pay_trucking_unit_price),
                "freight": _to_float(record.pay_trucking_freight),
                "doc_fee": _to_float(record.pay_trucking_doc_fee),
                "other_fee": _to_float(record.pay_trucking_other_fee),
                "remark": record.pay_trucking_remark or "",
            },
            "dom_air": {
                "subtotal": _to_float(record.pay_dom_air_subtotal),
                "date": _to_date_str(record.pay_dom_air_date),
                "outsource_unit": record.pay_dom_air_outsource_unit or "",
                "origin": record.pay_dom_air_origin or "",
                "destination": record.pay_dom_air_destination or "",
                "airline": record.pay_dom_air_airline or "",
                "airline_unit": record.pay_dom_air_airline_unit or "",
                "flight_doc_no": record.pay_dom_air_flight_doc_no or "",
                "flight_no": record.pay_dom_air_flight_no or "",
                "flight_date": _to_date_str(record.pay_dom_air_flight_date),
                "pieces": record.pay_dom_air_pieces,
                "weight": _to_float(record.pay_dom_air_weight),
                "chargeable_weight": _to_float(record.pay_dom_air_chargeable_weight),
                "rate": _to_float(record.pay_dom_air_rate),
                "freight": _to_float(record.pay_dom_air_freight),
                "other_fee": _to_float(record.pay_dom_air_other_fee),
                "remark": record.pay_dom_air_remark or "",
            },
            "customs": {
                "subtotal": _to_float(record.pay_customs_subtotal),
                "date": _to_date_str(record.pay_customs_date),
                "agent": record.pay_customs_agent or "",
                "customs_fee": _to_float(record.pay_customs_fee),
                "continuation_sheet_fee": _to_float(record.pay_customs_continuation_sheet_fee),
                "inspection_delete_fee": _to_float(record.pay_customs_inspection_delete_fee),
                "rebate": _to_float(record.pay_customs_rebate),
                "other_fee": _to_float(record.pay_customs_other_fee),
                "remark": record.pay_customs_remark or "",
            },
            "ground": {
                "subtotal": _to_float(record.pay_ground_subtotal),
                "date": _to_date_str(record.pay_ground_date),
                "outsource_unit": record.pay_ground_outsource_unit or "",
                "chargeable_weight": _to_float(record.pay_ground_chargeable_weight),
                "rate": _to_float(record.pay_ground_rate),
                "freight": _to_float(record.pay_ground_freight),
                "lading_express_fee": _to_float(record.pay_ground_lading_express_fee),
                "security_customs_fee": _to_float(record.pay_ground_security_customs_fee),
                "pallet_exit_fee": _to_float(record.pay_ground_pallet_exit_fee),
                "other_fee": _to_float(record.pay_ground_other_fee),
                "remark": record.pay_ground_remark or "",
            },
            "pay_total": _to_float(record.pay_total),
        },
        
        # (4) 销售提成
        "sales_commission": {
            "salesperson": record.salesperson or "",
            "commission_amount": _to_float(record.commission_amount),
        },
        
        # (5) 经营信息
        "operating_info": {
            "profit": _to_float(record.profit),
            "profit_margin": _to_float(record.profit_margin),
        },
        
        "created_at": format_datetime_china(record.created_at),
        "updated_at": format_datetime_china(record.updated_at),
    }
    if hasattr(record, "creator_id") and record.creator_id:
        data["creator_id"] = str(record.creator_id)
    return data


def _apply_cost_payload(record: Any, payload: CostRegistrationSave):
    """从层级化 Payload 赋值属性到 ORM 模型对象"""
    # 1. 货主委托信息
    if payload.consignor_info is not None:
        info = payload.consignor_info
        record.create_time = _parse_datetime(info.create_time) if info.create_time is not None else record.create_time
        record.internal_doc_id = info.internal_doc_id if info.internal_doc_id is not None else record.internal_doc_id
        record.warehouse_entry_date = _parse_date(info.warehouse_entry_date) if info.warehouse_entry_date is not None else record.warehouse_entry_date
        record.customer_name = info.customer_name if info.customer_name is not None else record.customer_name
        record.origin_destination = info.origin_destination if info.origin_destination is not None else record.origin_destination
        record.customs_declaration = info.customs_declaration if info.customs_declaration is not None else record.customs_declaration
        record.bill_of_lading = info.bill_of_lading if info.bill_of_lading is not None else record.bill_of_lading
        record.flight_date = _parse_date(info.flight_date) if info.flight_date is not None else record.flight_date
        record.flight_no = info.flight_no if info.flight_no is not None else record.flight_no
        record.flight_doc_no = info.flight_doc_no if info.flight_doc_no is not None else record.flight_doc_no
        record.pieces = info.pieces if info.pieces is not None else record.pieces
        record.actual_weight = info.actual_weight if info.actual_weight is not None else record.actual_weight
        record.chargeable_weight = info.chargeable_weight if info.chargeable_weight is not None else record.chargeable_weight
        record.volume = info.volume if info.volume is not None else record.volume
        record.first_leg_weight = info.first_leg_weight if info.first_leg_weight is not None else record.first_leg_weight
        record.agent = info.agent if info.agent is not None else record.agent
        record.remark = info.remark if info.remark is not None else record.remark

    # 2. 应收款项
    if payload.receivables is not None:
        rec = payload.receivables
        record.unit_price = rec.unit_price if rec.unit_price is not None else record.unit_price
        record.receivable_freight = rec.freight if rec.freight is not None else record.receivable_freight
        record.receivable_lading_info_fee = rec.lading_info_fee if rec.lading_info_fee is not None else record.receivable_lading_info_fee
        record.receivable_split_offset_telex_fee = rec.split_offset_telex_fee if rec.split_offset_telex_fee is not None else record.receivable_split_offset_telex_fee
        record.receivable_customs_fee = rec.customs_fee if rec.customs_fee is not None else record.receivable_customs_fee
        record.receivable_continuation_sheet_fee = rec.continuation_sheet_fee if rec.continuation_sheet_fee is not None else record.receivable_continuation_sheet_fee
        record.receivable_customs_inspection_fee = rec.customs_inspection_fee if rec.customs_inspection_fee is not None else record.receivable_customs_inspection_fee
        record.receivable_magnetic_security_fee = rec.magnetic_security_fee if rec.magnetic_security_fee is not None else record.receivable_magnetic_security_fee
        record.receivable_tc_express_fee = rec.tc_express_fee if rec.tc_express_fee is not None else record.receivable_tc_express_fee
        record.receivable_warehouse_ground_fee = rec.warehouse_ground_fee if rec.warehouse_ground_fee is not None else record.receivable_warehouse_ground_fee
        record.receivable_doc_make_fee = rec.doc_make_fee if rec.doc_make_fee is not None else record.receivable_doc_make_fee
        record.receivable_doc_split_fee = rec.doc_split_fee if rec.doc_split_fee is not None else record.receivable_doc_split_fee
        record.receivable_skid_fee = rec.skid_fee if rec.skid_fee is not None else record.receivable_skid_fee
        record.receivable_pallet_packing_fee = rec.pallet_packing_fee if rec.pallet_packing_fee is not None else record.receivable_pallet_packing_fee
        record.receivable_probe_fee = rec.probe_fee if rec.probe_fee is not None else record.receivable_probe_fee
        record.receivable_consumables_fee = rec.consumables_fee if rec.consumables_fee is not None else record.receivable_consumables_fee
        record.receivable_first_leg_fee = rec.first_leg_fee if rec.first_leg_fee is not None else record.receivable_first_leg_fee
        record.receivable_total = rec.total if rec.total is not None else record.receivable_total

    # 3. 应付款项
    if payload.payables is not None:
        p = payload.payables
        record.pay_total = p.pay_total if p.pay_total is not None else record.pay_total

        # [1] 国际空运
        if p.intl_air is not None:
            ia = p.intl_air
            record.pay_intl_air_subtotal = ia.subtotal if ia.subtotal is not None else record.pay_intl_air_subtotal
            record.pay_intl_air_date = _parse_date(ia.date) if ia.date is not None else record.pay_intl_air_date
            record.pay_intl_air_outsource_unit = ia.outsource_unit if ia.outsource_unit is not None else record.pay_intl_air_outsource_unit
            record.pay_intl_air_origin = ia.origin if ia.origin is not None else record.pay_intl_air_origin
            record.pay_intl_air_destination = ia.destination if ia.destination is not None else record.pay_intl_air_destination
            record.pay_intl_air_airline = ia.airline if ia.airline is not None else record.pay_intl_air_airline
            record.pay_intl_air_flight_doc_no = ia.flight_doc_no if ia.flight_doc_no is not None else record.pay_intl_air_flight_doc_no
            record.pay_intl_air_flight_no = ia.flight_no if ia.flight_no is not None else record.pay_intl_air_flight_no
            record.pay_intl_air_flight_date = _parse_date(ia.flight_date) if ia.flight_date is not None else record.pay_intl_air_flight_date
            record.pay_intl_air_pieces = ia.pieces if ia.pieces is not None else record.pay_intl_air_pieces
            record.pay_intl_air_weight = ia.weight if ia.weight is not None else record.pay_intl_air_weight
            record.pay_intl_air_volume = ia.volume if ia.volume is not None else record.pay_intl_air_volume
            record.pay_intl_air_chargeable_weight = ia.chargeable_weight if ia.chargeable_weight is not None else record.pay_intl_air_chargeable_weight
            record.pay_intl_air_rate = ia.rate if ia.rate is not None else record.pay_intl_air_rate
            record.pay_intl_air_freight = ia.freight if ia.freight is not None else record.pay_intl_air_freight
            record.pay_intl_air_lading_fee = ia.lading_fee if ia.lading_fee is not None else record.pay_intl_air_lading_fee
            record.pay_intl_air_split_fee = ia.split_fee if ia.split_fee is not None else record.pay_intl_air_split_fee
            record.pay_intl_air_borrow_magnetic_fuel_pickup_fee = ia.borrow_magnetic_fuel_pickup_fee if ia.borrow_magnetic_fuel_pickup_fee is not None else record.pay_intl_air_borrow_magnetic_fuel_pickup_fee
            record.pay_intl_air_tc_network_disposal_fee = ia.tc_network_disposal_fee if ia.tc_network_disposal_fee is not None else record.pay_intl_air_tc_network_disposal_fee
            record.pay_intl_air_customs_fee = ia.customs_fee if ia.customs_fee is not None else record.pay_intl_air_customs_fee
            record.pay_intl_air_continuation_sheet_fee = ia.continuation_sheet_fee if ia.continuation_sheet_fee is not None else record.pay_intl_air_continuation_sheet_fee
            record.pay_intl_air_consumables_fee = ia.consumables_fee if ia.consumables_fee is not None else record.pay_intl_air_consumables_fee
            record.pay_intl_air_front_warehouse = ia.front_warehouse if ia.front_warehouse is not None else record.pay_intl_air_front_warehouse
            record.pay_intl_air_other_fee = ia.other_fee if ia.other_fee is not None else record.pay_intl_air_other_fee
            record.pay_intl_air_remark = ia.remark if ia.remark is not None else record.pay_intl_air_remark

        # [2] 汽运
        if p.trucking is not None:
            tr = p.trucking
            record.pay_trucking_subtotal = tr.subtotal if tr.subtotal is not None else record.pay_trucking_subtotal
            record.pay_trucking_date = _parse_date(tr.date) if tr.date is not None else record.pay_trucking_date
            record.pay_trucking_outsource_unit = tr.outsource_unit if tr.outsource_unit is not None else record.pay_trucking_outsource_unit
            record.pay_trucking_pieces = tr.pieces if tr.pieces is not None else record.pay_trucking_pieces
            record.pay_trucking_weight = tr.weight if tr.weight is not None else record.pay_trucking_weight
            record.pay_trucking_volume = tr.volume if tr.volume is not None else record.pay_trucking_volume
            record.pay_trucking_unit_price = tr.unit_price if tr.unit_price is not None else record.pay_trucking_unit_price
            record.pay_trucking_freight = tr.freight if tr.freight is not None else record.pay_trucking_freight
            record.pay_trucking_doc_fee = tr.doc_fee if tr.doc_fee is not None else record.pay_trucking_doc_fee
            record.pay_trucking_other_fee = tr.other_fee if tr.other_fee is not None else record.pay_trucking_other_fee
            record.pay_trucking_remark = tr.remark if tr.remark is not None else record.pay_trucking_remark

        # [3] 国内空运
        if p.dom_air is not None:
            da = p.dom_air
            record.pay_dom_air_subtotal = da.subtotal if da.subtotal is not None else record.pay_dom_air_subtotal
            record.pay_dom_air_date = _parse_date(da.date) if da.date is not None else record.pay_dom_air_date
            record.pay_dom_air_outsource_unit = da.outsource_unit if da.outsource_unit is not None else record.pay_dom_air_outsource_unit
            record.pay_dom_air_origin = da.origin if da.origin is not None else record.pay_dom_air_origin
            record.pay_dom_air_destination = da.destination if da.destination is not None else record.pay_dom_air_destination
            record.pay_dom_air_airline = da.airline if da.airline is not None else record.pay_dom_air_airline
            record.pay_dom_air_airline_unit = da.airline_unit if da.airline_unit is not None else record.pay_dom_air_airline_unit
            record.pay_dom_air_flight_doc_no = da.flight_doc_no if da.flight_doc_no is not None else record.pay_dom_air_flight_doc_no
            record.pay_dom_air_flight_no = da.flight_no if da.flight_no is not None else record.pay_dom_air_flight_no
            record.pay_dom_air_flight_date = _parse_date(da.flight_date) if da.flight_date is not None else record.pay_dom_air_flight_date
            record.pay_dom_air_pieces = da.pieces if da.pieces is not None else record.pay_dom_air_pieces
            record.pay_dom_air_weight = da.weight if da.weight is not None else record.pay_dom_air_weight
            record.pay_dom_air_chargeable_weight = da.chargeable_weight if da.chargeable_weight is not None else record.pay_dom_air_chargeable_weight
            record.pay_dom_air_rate = da.rate if da.rate is not None else record.pay_dom_air_rate
            record.pay_dom_air_freight = da.freight if da.freight is not None else record.pay_dom_air_freight
            record.pay_dom_air_other_fee = da.other_fee if da.other_fee is not None else record.pay_dom_air_other_fee
            record.pay_dom_air_remark = da.remark if da.remark is not None else record.pay_dom_air_remark

        # [4] 报关信息
        if p.customs is not None:
            c = p.customs
            record.pay_customs_subtotal = c.subtotal if c.subtotal is not None else record.pay_customs_subtotal
            record.pay_customs_date = _parse_date(c.date) if c.date is not None else record.pay_customs_date
            record.pay_customs_agent = c.agent if c.agent is not None else record.pay_customs_agent
            record.pay_customs_fee = c.customs_fee if c.customs_fee is not None else record.pay_customs_fee
            record.pay_customs_continuation_sheet_fee = c.continuation_sheet_fee if c.continuation_sheet_fee is not None else record.pay_customs_continuation_sheet_fee
            record.pay_customs_inspection_delete_fee = c.inspection_delete_fee if c.inspection_delete_fee is not None else record.pay_customs_inspection_delete_fee
            record.pay_customs_rebate = c.rebate if c.rebate is not None else record.pay_customs_rebate
            record.pay_customs_other_fee = c.other_fee if c.other_fee is not None else record.pay_customs_other_fee
            record.pay_customs_remark = c.remark if c.remark is not None else record.pay_customs_remark

        # [5] 地面操作信息
        if p.ground is not None:
            g = p.ground
            record.pay_ground_subtotal = g.subtotal if g.subtotal is not None else record.pay_ground_subtotal
            record.pay_ground_date = _parse_date(g.date) if g.date is not None else record.pay_ground_date
            record.pay_ground_outsource_unit = g.outsource_unit if g.outsource_unit is not None else record.pay_ground_outsource_unit
            record.pay_ground_chargeable_weight = g.chargeable_weight if g.chargeable_weight is not None else record.pay_ground_chargeable_weight
            record.pay_ground_rate = g.rate if g.rate is not None else record.pay_ground_rate
            record.pay_ground_freight = g.freight if g.freight is not None else record.pay_ground_freight
            record.pay_ground_lading_express_fee = g.lading_express_fee if g.lading_express_fee is not None else record.pay_ground_lading_express_fee
            record.pay_ground_security_customs_fee = g.security_customs_fee if g.security_customs_fee is not None else record.pay_ground_security_customs_fee
            record.pay_ground_pallet_exit_fee = g.pallet_exit_fee if g.pallet_exit_fee is not None else record.pay_ground_pallet_exit_fee
            record.pay_ground_other_fee = g.other_fee if g.other_fee is not None else record.pay_ground_other_fee
            record.pay_ground_remark = g.remark if g.remark is not None else record.pay_ground_remark

    # 4. 销售提成
    if payload.sales_commission is not None:
        sc = payload.sales_commission
        record.salesperson = sc.salesperson if sc.salesperson is not None else record.salesperson
        record.commission_amount = sc.commission_amount if sc.commission_amount is not None else record.commission_amount

    # 5. 经营信息
    if payload.operating_info is not None:
        op = payload.operating_info
        record.profit = op.profit if op.profit is not None else record.profit
        record.profit_margin = op.profit_margin if op.profit_margin is not None else record.profit_margin


# ============================================================================
# 1. 费用信息登记接口（系统唯一一条数据，支持编辑和保存）
# ============================================================================

@router.get("/cost-registration", summary="获取费用信息登记数据（系统唯一数据）")
async def get_cost_registration(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    获取系统当前维护的唯一一条费用信息登记模版数据。
    若系统尚未保存过该数据，则 data 返回 null。
    """
    record = db.query(CostRegistration).first()
    if not record:
        return success_response(data=None, msg="暂无费用信息登记数据")
    
    return success_response(data=_format_cost_record(record), msg="查询成功")


@router.put("/cost-registration", summary="编辑并保存费用信息登记数据（系统唯一数据）")
async def save_cost_registration(
    payload: CostRegistrationSave,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    保存/编辑费用信息登记接口（Upsert 操作）。
    系统整体仅维护一条数据，若不存在则新建，若存在则更新。
    """
    record = db.query(CostRegistration).first()
    
    if record:
        _apply_cost_payload(record, payload)
        db.commit()
        db.refresh(record)
        msg = "费用信息登记更新成功"
    else:
        record = CostRegistration()
        _apply_cost_payload(record, payload)
        db.add(record)
        db.commit()
        db.refresh(record)
        msg = "费用信息登记保存成功"

    return success_response(data=_format_cost_record(record), msg=msg)


# ============================================================================
# 2. 单据信息-新增
# ============================================================================

@router.post("/consignments", summary="单据信息-新增")
async def create_cost_consignment(
    payload: CostConsignmentCreate,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    新增一条费用单据记录。
    """
    new_record = CostConsignment(creator_id=current_user.id)
    _apply_cost_payload(new_record, payload)
    
    # 若制单时间未传入，自动填充当前时间
    if not new_record.create_time:
        new_record.create_time = get_china_now()

    db.add(new_record)
    db.commit()
    db.refresh(new_record)
    
    return success_response(data=_format_cost_record(new_record), msg="单据信息创建成功")


# ============================================================================
# 3. 单据信息-列表（支持进仓日期区间、客户名称、代理单位、航司单号、航班号等条件查询及分页）
# ============================================================================

@router.get("/consignments", summary="单据信息-列表")
async def get_cost_consignments(
    start_warehouse_date: Optional[str] = Query(None, description="进仓开始日期 (YYYY-MM-DD)"),
    end_warehouse_date: Optional[str] = Query(None, description="进仓结束日期 (YYYY-MM-DD)"),
    customer_name: Optional[str] = Query(None, description="客户名称 (模糊查询)"),
    agent: Optional[str] = Query(None, description="代理单位 (模糊查询)"),
    flight_doc_no: Optional[str] = Query(None, description="航司单号/航班单号 (模糊查询)"),
    flight_no: Optional[str] = Query(None, description="航班号 (模糊查询)"),
    page: Optional[int] = Query(1, ge=1, description="页码"),
    pageSize: Optional[int] = Query(10, ge=1, description="每页数量"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    单据信息列表查询接口。
    
    支持参数：
    - **start_warehouse_date**: 进仓日期区间开始，例如 '2026-07-25'
    - **end_warehouse_date**: 进仓日期区间结束，例如 '2026-07-30'
    - **customer_name**: 客户名称 (支持模糊匹配)
    - **agent**: 代理单位 (支持模糊匹配)
    - **flight_doc_no**: 航司单号/航班单号 (支持模糊匹配)
    - **flight_no**: 航班号 (支持模糊匹配)
    - **page**: 页码
    - **pageSize**: 每页条数
    """
    query_obj = db.query(CostConsignment)
    
    # 1. 进仓日期区间筛选
    if start_warehouse_date:
        s_date = _parse_date(start_warehouse_date)
        if s_date:
            query_obj = query_obj.filter(CostConsignment.warehouse_entry_date >= s_date)
            
    if end_warehouse_date:
        e_date = _parse_date(end_warehouse_date)
        if e_date:
            query_obj = query_obj.filter(CostConsignment.warehouse_entry_date <= e_date)
            
    # 2. 客户名称模糊查询
    if customer_name and customer_name.strip():
        query_obj = query_obj.filter(CostConsignment.customer_name.like(f"%{customer_name.strip()}%"))
        
    # 3. 代理单位模糊查询
    if agent and agent.strip():
        query_obj = query_obj.filter(CostConsignment.agent.like(f"%{agent.strip()}%"))
        
    # 4. 航司单号/航班单号模糊查询
    if flight_doc_no and flight_doc_no.strip():
        doc_no = flight_doc_no.strip()
        query_obj = query_obj.filter(
            (CostConsignment.flight_doc_no.like(f"%{doc_no}%")) |
            (CostConsignment.bill_of_lading.like(f"%{doc_no}%")) |
            (CostConsignment.pay_intl_air_flight_doc_no.like(f"%{doc_no}%")) |
            (CostConsignment.pay_dom_air_flight_doc_no.like(f"%{doc_no}%"))
        )
        
    # 5. 航班号模糊查询
    if flight_no and flight_no.strip():
        f_no = flight_no.strip()
        query_obj = query_obj.filter(
            (CostConsignment.flight_no.like(f"%{f_no}%")) |
            (CostConsignment.pay_intl_air_flight_no.like(f"%{f_no}%")) |
            (CostConsignment.pay_dom_air_flight_no.like(f"%{f_no}%"))
        )
            
    total = query_obj.count()
    
    # 排序：进仓日期倒序，其次制单时间倒序，其次ID倒序
    query_obj = query_obj.order_by(
        CostConsignment.warehouse_entry_date.desc(),
        CostConsignment.create_time.desc(),
        CostConsignment.id.desc()
    )
    
    # 分页
    if page is not None and pageSize is not None:
        offset = (page - 1) * pageSize
        records = query_obj.offset(offset).limit(pageSize).all()
    else:
        records = query_obj.all()
        
    items = [_format_cost_record(r) for r in records]
    
    return success_response(
        data={"total": total, "items": items, "page": page, "pageSize": pageSize},
        msg="查询成功"
    )


# ============================================================================
# 单据信息-详情
# ============================================================================

@router.get("/consignments/{consignment_id}", summary="单据信息-详情")
async def get_cost_consignment_detail(
    consignment_id: str = Path(..., description="单据明细ID"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """获取单条费用单据详情接口"""
    try:
        c_id = int(consignment_id)
    except ValueError:
        raise BadRequestException("consignment_id 必须为合法数字格式")
        
    record = db.query(CostConsignment).filter(CostConsignment.id == c_id).first()
    if not record:
        raise NotFoundException(f"单据信息不存在 (ID: {consignment_id})")
        
    return success_response(data=_format_cost_record(record), msg="查询成功")


# ============================================================================
# 5. 单据信息-修改
# ============================================================================

@router.put("/consignments/{consignment_id}", summary="单据信息-修改")
async def update_cost_consignment(
    payload: CostConsignmentUpdate,
    consignment_id: str = Path(..., description="单据明细ID"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    修改指定 ID 的费用单据信息接口。
    """
    try:
        c_id = int(consignment_id)
    except ValueError:
        raise BadRequestException("consignment_id 必须为合法数字格式")
        
    record = db.query(CostConsignment).filter(CostConsignment.id == c_id).first()
    if not record:
        raise NotFoundException(f"单据信息不存在 (ID: {consignment_id})")
        
    _apply_cost_payload(record, payload)
    db.commit()
    db.refresh(record)
    
    return success_response(data=_format_cost_record(record), msg="单据信息更新成功")


# ============================================================================
# 4. 单据信息-删除 (批量删除与单个删除)
# ============================================================================

@router.post("/consignments/batch-delete", summary="单据信息-批量删除 (POST)")
@router.delete("/consignments/batch-delete", summary="单据信息-批量删除 (DELETE)")
async def batch_delete_cost_consignments(
    payload: CostBatchDeleteRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    批量删除费用单据接口（同时支持 POST 与 DELETE 方法）。
    
    传入 JSON 数组：`{"ids": ["123", "456"]}`
    """
    if not payload.ids:
        raise BadRequestException("待删除的 ID 数组不能为空")
        
    int_ids = []
    for raw_id in payload.ids:
        try:
            int_ids.append(int(raw_id))
        except ValueError:
            raise BadRequestException(f"ID '{raw_id}' 格式无效")
            
    deleted_count = db.query(CostConsignment).filter(CostConsignment.id.in_(int_ids)).delete(synchronize_session=False)
    db.commit()
    
    return success_response(
        data={"deleted_count": deleted_count, "requested_ids": payload.ids},
        msg=f"成功批量删除 {deleted_count} 条单据信息记录"
    )


@router.delete("/consignments/{consignment_id}", summary="单据信息-删除（单个）")
async def delete_cost_consignment(
    consignment_id: str = Path(..., description="单据明细ID"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """单个删除费用单据接口"""
    try:
        c_id = int(consignment_id)
    except ValueError:
        raise BadRequestException("consignment_id 必须为合法数字格式")
        
    record = db.query(CostConsignment).filter(CostConsignment.id == c_id).first()
    if not record:
        raise NotFoundException(f"单据信息不存在 (ID: {consignment_id})")
        
    db.delete(record)
    db.commit()
    
    return success_response(data={"id": consignment_id}, msg="单据信息删除成功")


# ============================================================================
# 6. 单据信息-选中下载为 excel
# ============================================================================

@router.post("/consignments/export-excel", summary="单据信息-选中下载为excel")
async def export_cost_consignments_to_excel(
    payload: CostExportExcelRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    选中费用单据列表中的某些项导出为 Excel (.xlsx) 表格文件。
    涵盖 5 大业务层级结构（货主委托、应收明细、应付明细[国空/汽运/国内/报关/地面]、销售提成、经营信息）共 113 列全量字段。
    
    传入选中的 ID 数组：`{"ids": ["123", "456"]}`
    """
    if not payload.ids:
        raise BadRequestException("请选择至少一条需导出的单据记录")
        
    int_ids = []
    for raw_id in payload.ids:
        try:
            int_ids.append(int(raw_id))
        except ValueError:
            raise BadRequestException(f"ID '{raw_id}' 格式无效")
            
    records = db.query(CostConsignment).filter(CostConsignment.id.in_(int_ids)).order_by(CostConsignment.warehouse_entry_date.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "费用单据信息全量列表"
    
    headers = [
        # (1) 货主委托信息
        "制单时间", "内部单据ID", "进仓日期", "客户名称", "始发站-目的站",
        "报关", "提单", "航班日期", "航班号", "航班单号",
        "件数", "实际重量(kg)", "计费重量(kg)", "体积(m³)", "一程重量(kg)",
        "代理", "委托备注",
        
        # (2) 应收款项
        "应收-单价", "应收-运费", "应收-提单费/信息录入费", "应收-分单费/抵账费/电报费",
        "应收-报关费", "应收-续页费", "应收-海关查验费", "应收-磁检费/安检费",
        "应收-TC操作费/快件中心过站费", "应收-前置仓/国际货站地面费", "应收-制单费",
        "应收-制单分单费", "应收-垫板费", "应收-打板/装箱费", "应收-探板费",
        "应收-耗材费", "应收-一程费用", "应收-合计",
        
        # (3) 应付款项 - 国际空运
        "国空应付-小计", "国空应付-托运日期", "国空应付-外发单位", "国空应付-始发站",
        "国空应付-到达站", "国空应付-航空公司", "国空应付-航班单号", "国空应付-航班号",
        "国空应付-航班日期", "国空应付-件数", "国空应付-重量", "国空应付-体积",
        "国空应付-计费重量", "国空应付-费率", "国空应付-运费", "国空应付-提单费",
        "国空应付-分单费", "国空应付-借单/磁检/燃油/提货费", "国空应付-TC/入网/处置费",
        "国空应付-报关费", "国空应付-续页费", "国空应付-耗材费", "国空应付-前置仓",
        "国空应付-其他费用", "国空应付-备注",
        
        # (3) 应付款项 - 汽运
        "汽运应付-小计", "汽运应付-托运日期", "汽运应付-外发单位", "汽运应付-件数",
        "汽运应付-重量", "汽运应付-体积", "汽运应付-单价", "汽运应付-运费",
        "汽运应付-制单费", "汽运应付-其他费用", "汽运应付-备注",
        
        # (3) 应付款项 - 国内空运
        "国空内应付-小计", "国空内应付-托运日期", "国空内应付-外发单位", "国空内应付-始发站",
        "国空内应付-到达站", "国空内应付-航空公司", "国空内应付-航空单位", "国空内应付-航空单号",
        "国空内应付-航班号", "国空内应付-航班日期", "国空内应付-件数", "国空内应付-重量",
        "国空内应付-计费重量", "国空内应付-费率", "国空内应付-运费", "国空内应付-其他费用", "国空内应付-备注",
        
        # (3) 应付款项 - 报关
        "报关应付-小计", "报关应付-报关日期", "报关应付-报关代理", "报关应付-报关费",
        "报关应付-续页费", "报关应付-查验/删单费", "报关应付-回扣栏", "报关应付-其他费用", "报关应付-备注",
        
        # (3) 应付款项 - 地面操作
        "地面应付-小计", "地面应付-托运日期", "地面应付-外发单位", "地面应付-计费重量",
        "地面应付-费率", "地面应付-运费", "地面应付-提单/快件处置费", "地面应付-安检/报关费",
        "地面应付-打板/退场费", "地面应付-其他费用", "地面应付-备注",
        
        # (3) 应付款项 - 总计
        "应付合计",
        
        # (4) 销售提成
        "业务员", "提成金额",
        
        # (5) 经营信息
        "利润", "利润率(%)"
    ]
    
    ws.append(headers)
    
    # 样式配置
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="微软雅黑", size=9)
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 28
    
    def _v_str(val):
        return str(val) if val is not None else ""

    def _v_num(val):
        return float(val) if val is not None else ""

    def _v_date(val):
        if not val:
            return ""
        if isinstance(val, (datetime, date)):
            return val.strftime("%Y-%m-%d")
        return str(val)

    def _v_dt(val):
        if not val:
            return ""
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d %H:%M:%S")
        return str(val)

    for r_idx, rec in enumerate(records, 2):
        row_data = [
            # (1) 货主委托信息
            _v_dt(rec.create_time),
            _v_str(rec.internal_doc_id),
            _v_date(rec.warehouse_entry_date),
            _v_str(rec.customer_name),
            _v_str(rec.origin_destination),
            _v_str(rec.customs_declaration),
            _v_str(rec.bill_of_lading),
            _v_date(rec.flight_date),
            _v_str(rec.flight_no),
            _v_str(rec.flight_doc_no),
            rec.pieces if rec.pieces is not None else "",
            _v_num(rec.actual_weight),
            _v_num(rec.chargeable_weight),
            _v_num(rec.volume),
            _v_num(rec.first_leg_weight),
            _v_str(rec.agent),
            _v_str(rec.remark),

            # (2) 应收款项
            _v_num(rec.unit_price),
            _v_num(rec.receivable_freight),
            _v_num(rec.receivable_lading_info_fee),
            _v_num(rec.receivable_split_offset_telex_fee),
            _v_num(rec.receivable_customs_fee),
            _v_num(rec.receivable_continuation_sheet_fee),
            _v_num(rec.receivable_customs_inspection_fee),
            _v_num(rec.receivable_magnetic_security_fee),
            _v_num(rec.receivable_tc_express_fee),
            _v_num(rec.receivable_warehouse_ground_fee),
            _v_num(rec.receivable_doc_make_fee),
            _v_num(rec.receivable_doc_split_fee),
            _v_num(rec.receivable_skid_fee),
            _v_num(rec.receivable_pallet_packing_fee),
            _v_num(rec.receivable_probe_fee),
            _v_num(rec.receivable_consumables_fee),
            _v_num(rec.receivable_first_leg_fee),
            _v_num(rec.receivable_total),

            # (3) 应付款项 - 国际空运
            _v_num(rec.pay_intl_air_subtotal),
            _v_date(rec.pay_intl_air_date),
            _v_str(rec.pay_intl_air_outsource_unit),
            _v_str(rec.pay_intl_air_origin),
            _v_str(rec.pay_intl_air_destination),
            _v_str(rec.pay_intl_air_airline),
            _v_str(rec.pay_intl_air_flight_doc_no),
            _v_str(rec.pay_intl_air_flight_no),
            _v_date(rec.pay_intl_air_flight_date),
            rec.pay_intl_air_pieces if rec.pay_intl_air_pieces is not None else "",
            _v_num(rec.pay_intl_air_weight),
            _v_num(rec.pay_intl_air_volume),
            _v_num(rec.pay_intl_air_chargeable_weight),
            _v_num(rec.pay_intl_air_rate),
            _v_num(rec.pay_intl_air_freight),
            _v_num(rec.pay_intl_air_lading_fee),
            _v_num(rec.pay_intl_air_split_fee),
            _v_num(rec.pay_intl_air_borrow_magnetic_fuel_pickup_fee),
            _v_num(rec.pay_intl_air_tc_network_disposal_fee),
            _v_num(rec.pay_intl_air_customs_fee),
            _v_num(rec.pay_intl_air_continuation_sheet_fee),
            _v_num(rec.pay_intl_air_consumables_fee),
            _v_num(rec.pay_intl_air_front_warehouse),
            _v_num(rec.pay_intl_air_other_fee),
            _v_str(rec.pay_intl_air_remark),

            # (3) 应付款项 - 汽运
            _v_num(rec.pay_trucking_subtotal),
            _v_date(rec.pay_trucking_date),
            _v_str(rec.pay_trucking_outsource_unit),
            rec.pay_trucking_pieces if rec.pay_trucking_pieces is not None else "",
            _v_num(rec.pay_trucking_weight),
            _v_num(rec.pay_trucking_volume),
            _v_num(rec.pay_trucking_unit_price),
            _v_num(rec.pay_trucking_freight),
            _v_num(rec.pay_trucking_doc_fee),
            _v_num(rec.pay_trucking_other_fee),
            _v_str(rec.pay_trucking_remark),

            # (3) 应付款项 - 国内空运
            _v_num(rec.pay_dom_air_subtotal),
            _v_date(rec.pay_dom_air_date),
            _v_str(rec.pay_dom_air_outsource_unit),
            _v_str(rec.pay_dom_air_origin),
            _v_str(rec.pay_dom_air_destination),
            _v_str(rec.pay_dom_air_airline),
            _v_str(rec.pay_dom_air_airline_unit),
            _v_str(rec.pay_dom_air_flight_doc_no),
            _v_str(rec.pay_dom_air_flight_no),
            _v_date(rec.pay_dom_air_flight_date),
            rec.pay_dom_air_pieces if rec.pay_dom_air_pieces is not None else "",
            _v_num(rec.pay_dom_air_weight),
            _v_num(rec.pay_dom_air_chargeable_weight),
            _v_num(rec.pay_dom_air_rate),
            _v_num(rec.pay_dom_air_freight),
            _v_num(rec.pay_dom_air_other_fee),
            _v_str(rec.pay_dom_air_remark),

            # (3) 应付款项 - 报关
            _v_num(rec.pay_customs_subtotal),
            _v_date(rec.pay_customs_date),
            _v_str(rec.pay_customs_agent),
            _v_num(rec.pay_customs_fee),
            _v_num(rec.pay_customs_continuation_sheet_fee),
            _v_num(rec.pay_customs_inspection_delete_fee),
            _v_num(rec.pay_customs_rebate),
            _v_num(rec.pay_customs_other_fee),
            _v_str(rec.pay_customs_remark),

            # (3) 应付款项 - 地面操作
            _v_num(rec.pay_ground_subtotal),
            _v_date(rec.pay_ground_date),
            _v_str(rec.pay_ground_outsource_unit),
            _v_num(rec.pay_ground_chargeable_weight),
            _v_num(rec.pay_ground_rate),
            _v_num(rec.pay_ground_freight),
            _v_num(rec.pay_ground_lading_express_fee),
            _v_num(rec.pay_ground_security_customs_fee),
            _v_num(rec.pay_ground_pallet_exit_fee),
            _v_num(rec.pay_ground_other_fee),
            _v_str(rec.pay_ground_remark),

            # (3) 应付款项 - 总计
            _v_num(rec.pay_total),

            # (4) 销售提成
            _v_str(rec.salesperson),
            _v_num(rec.commission_amount),

            # (5) 经营信息
            _v_num(rec.profit),
            _v_num(rec.profit_margin),
        ]
        
        ws.append(row_data)
        ws.row_dimensions[r_idx].height = 22
        
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.border = thin_border
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif any(k in headers[c_idx - 1] for k in ("时间", "日期")):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # 自动自适应列宽
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            length = sum(2 if ord(char) > 127 else 1 for char in val_str)
            if length > max_len:
                max_len = length
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"cost_consignment_export_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )
