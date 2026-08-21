"""
客服接单台 API 接口
"""
import io
from datetime import datetime, date
from typing import List, Optional, Any, Dict

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from fastapi import APIRouter, Depends, Path, Query
from fastapi.responses import Response
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.database import get_db
from app.core.exceptions import NotFoundException, BadRequestException
from app.core.response import success_response
from app.api.deps import get_current_active_user
from app.models.user import User
from app.models.customer_service import ConsignmentRegistration, ConsignmentInfo
from app.models.cost_service import CostConsignment
from app.schemas.customer_service import (
    ConsignmentRegistrationSave,
    ConsignmentRegistrationResponse,
    ConsignmentInfoCreate,
    ConsignmentInfoUpdate,
    ConsignmentInfoQuery,
    ConsignmentInfoSortField,
    ConsignmentInfoSortOrder,
    BatchDeleteRequest,
    ExportExcelRequest,
    ConsignmentInfoResponse,
)
from app.services.cost_excel_export import format_bill_of_lading_for_export
from app.utils.helpers import format_datetime_china, get_china_now

router = APIRouter()


def _parse_datetime(val: Any) -> Optional[datetime]:
    """解析日期时间"""
    if not val:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        val_str = val.strip()
        if not val_str:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(val_str, fmt)
            except ValueError:
                continue
    return None


def _parse_date(val: Any) -> Optional[date]:
    """解析日期"""
    if not val:
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, str):
        val_str = val.strip()
        if not val_str:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(val_str, fmt).date()
            except ValueError:
                continue
    return None


def _format_record_dict(record: Any) -> Dict[str, Any]:
    """将 SQLAlchemy 记录格式化为字典响应格式"""
    if not record:
        return {}
    
    create_time_str = None
    if record.create_time:
        create_time_str = record.create_time.strftime("%Y-%m-%d %H:%M:%S")
        
    warehouse_entry_date_str = None
    if record.warehouse_entry_date:
        warehouse_entry_date_str = record.warehouse_entry_date.strftime("%Y-%m-%d")
        
    flight_date_str = None
    if record.flight_date:
        flight_date_str = record.flight_date.strftime("%Y-%m-%d")

    data = {
        "id": str(record.id),
        "create_time": create_time_str,
        "internal_doc_id": record.internal_doc_id or "",
        "warehouse_entry_date": warehouse_entry_date_str,
        "customer_name": record.customer_name or "",
        "origin_destination": record.origin_destination or "",
        "customs_declaration": record.customs_declaration or "",
        "bill_of_lading": record.bill_of_lading or "",
        "flight_date": flight_date_str,
        "flight_no": record.flight_no or "",
        "flight_doc_no": record.flight_doc_no or "",
        "pieces": record.pieces,
        "actual_weight": float(record.actual_weight) if record.actual_weight is not None else None,
        "chargeable_weight": float(record.chargeable_weight) if record.chargeable_weight is not None else None,
        "volume": float(record.volume) if record.volume is not None else None,
        "first_leg_weight": float(record.first_leg_weight) if record.first_leg_weight is not None else None,
        "agent": record.agent or "",
        "remark": record.remark or "",
        "created_at": format_datetime_china(record.created_at),
        "updated_at": format_datetime_china(record.updated_at),
    }
    if hasattr(record, "creator_id") and record.creator_id:
        data["creator_id"] = str(record.creator_id)
    return data


# ============================================================================
# 1. 委托信息登记接口（系统唯一一条数据，支持编辑和保存）
# ============================================================================

@router.get("/consignment-registration", summary="获取委托信息登记数据（系统唯一数据）")
async def get_consignment_registration(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    获取系统当前维护的唯一一条委托信息登记数据。
    若系统尚未保存过该数据，则 data 返回 null。
    """
    record = db.query(ConsignmentRegistration).first()
    if not record:
        return success_response(data=None, msg="暂无委托信息登记数据")
    
    return success_response(data=_format_record_dict(record), msg="查询成功")


@router.put("/consignment-registration", summary="编辑并保存委托信息登记数据（系统唯一数据）")
async def save_consignment_registration(
    payload: ConsignmentRegistrationSave,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    保存/编辑委托信息登记接口（Upsert 操作）。
    系统整体仅维护一条数据，若不存在则新建，若存在则更新。
    """
    record = db.query(ConsignmentRegistration).first()
    
    create_time_val = _parse_datetime(payload.create_time) or get_china_now()
    warehouse_entry_date_val = _parse_date(payload.warehouse_entry_date)
    flight_date_val = _parse_date(payload.flight_date)
    
    if record:
        record.create_time = create_time_val
        record.internal_doc_id = payload.internal_doc_id
        record.warehouse_entry_date = warehouse_entry_date_val
        record.customer_name = payload.customer_name
        record.origin_destination = payload.origin_destination
        record.customs_declaration = payload.customs_declaration
        record.bill_of_lading = payload.bill_of_lading
        record.flight_date = flight_date_val
        record.flight_no = payload.flight_no
        record.flight_doc_no = payload.flight_doc_no
        record.pieces = payload.pieces
        record.actual_weight = payload.actual_weight
        record.chargeable_weight = payload.chargeable_weight
        record.volume = payload.volume
        record.first_leg_weight = payload.first_leg_weight
        record.agent = payload.agent
        record.remark = payload.remark
        
        db.commit()
        db.refresh(record)
        msg = "委托信息登记更新成功"
    else:
        record = ConsignmentRegistration(
            create_time=create_time_val,
            internal_doc_id=payload.internal_doc_id,
            warehouse_entry_date=warehouse_entry_date_val,
            customer_name=payload.customer_name,
            origin_destination=payload.origin_destination,
            customs_declaration=payload.customs_declaration,
            bill_of_lading=payload.bill_of_lading,
            flight_date=flight_date_val,
            flight_no=payload.flight_no,
            flight_doc_no=payload.flight_doc_no,
            pieces=payload.pieces,
            actual_weight=payload.actual_weight,
            chargeable_weight=payload.chargeable_weight,
            volume=payload.volume,
            first_leg_weight=payload.first_leg_weight,
            agent=payload.agent,
            remark=payload.remark,
        )
        db.add(record)
        db.commit()
        db.refresh(record)
        msg = "委托信息登记保存成功"

    return success_response(data=_format_record_dict(record), msg=msg)


# ============================================================================
# 2. 委托信息-新增
# ============================================================================

@router.post("/consignments", summary="委托信息-新增")
async def create_consignment(
    payload: ConsignmentInfoCreate,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    新增一条委托信息记录。
    """
    create_time_val = _parse_datetime(payload.create_time) or get_china_now()
    warehouse_entry_date_val = _parse_date(payload.warehouse_entry_date)
    flight_date_val = _parse_date(payload.flight_date)
    
    new_record = ConsignmentInfo(
        create_time=create_time_val,
        internal_doc_id=payload.internal_doc_id,
        warehouse_entry_date=warehouse_entry_date_val,
        customer_name=payload.customer_name,
        origin_destination=payload.origin_destination,
        customs_declaration=payload.customs_declaration,
        bill_of_lading=payload.bill_of_lading,
        flight_date=flight_date_val,
        flight_no=payload.flight_no,
        flight_doc_no=payload.flight_doc_no,
        pieces=payload.pieces,
        actual_weight=payload.actual_weight,
        chargeable_weight=payload.chargeable_weight,
        volume=payload.volume,
        first_leg_weight=payload.first_leg_weight,
        agent=payload.agent,
        remark=payload.remark,
        creator_id=current_user.id
    )
    
    db.add(new_record)
    db.flush()
    
    # 同步在费用登记台 (CostConsignment) 中创建记录
    cost_record = CostConsignment(
        id=new_record.id,
        create_time=new_record.create_time,
        internal_doc_id=new_record.internal_doc_id,
        warehouse_entry_date=new_record.warehouse_entry_date,
        customer_name=new_record.customer_name,
        origin_destination=new_record.origin_destination,
        customs_declaration=new_record.customs_declaration,
        bill_of_lading=new_record.bill_of_lading,
        flight_date=new_record.flight_date,
        flight_no=new_record.flight_no,
        flight_doc_no=new_record.flight_doc_no,
        pieces=new_record.pieces,
        actual_weight=new_record.actual_weight,
        chargeable_weight=new_record.chargeable_weight,
        volume=new_record.volume,
        first_leg_weight=new_record.first_leg_weight,
        agent=new_record.agent,
        remark=new_record.remark,
        creator_id=current_user.id
    )
    db.add(cost_record)
    db.commit()
    db.refresh(new_record)
    
    return success_response(data=_format_record_dict(new_record), msg="委托信息创建成功")


# ============================================================================
# 3. 委托信息-列表（支持筛选、排序及分页）
# ============================================================================

@router.get("/consignments", summary="委托信息-列表")
async def get_consignments(
    start_date: Optional[str] = Query(None, description="制单开始日期 (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="制单结束日期 (YYYY-MM-DD)"),
    customer_name: Optional[str] = Query(None, description="客户名称 (模糊查询)"),
    sort_by: ConsignmentInfoSortField = Query(
        ConsignmentInfoSortField.CREATE_TIME,
        description="排序字段：create_time（制单时间）或 warehouse_entry_date（进仓日期）",
    ),
    sort_order: ConsignmentInfoSortOrder = Query(
        ConsignmentInfoSortOrder.DESC,
        description="排序方向：asc（正序）或 desc（倒序）",
    ),
    page: Optional[int] = Query(1, ge=1, description="页码"),
    pageSize: Optional[int] = Query(10, ge=1, description="每页数量"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    委托信息列表查询接口。
    
    支持参数：
    - **start_date**: 制单日期区间开始，例如 '2026-07-25'
    - **end_date**: 制单日期区间结束，例如 '2026-07-30'
    - **customer_name**: 客户名称 (支持模糊匹配)
    - **sort_by**: 排序字段，可选 `create_time` 或 `warehouse_entry_date`，默认 `create_time`
    - **sort_order**: 排序方向，可选 `asc` 或 `desc`，默认 `desc`
    - **page**: 页码（不传或默认为 1）
    - **pageSize**: 每页条数（不传或默认为 10）
    """
    query_obj = db.query(ConsignmentInfo)
    
    # 日期区间筛选 (基于 create_time 字段)
    if start_date:
        s_date = _parse_date(start_date)
        if s_date:
            s_dt = datetime.combine(s_date, datetime.min.time())
            query_obj = query_obj.filter(ConsignmentInfo.create_time >= s_dt)
            
    if end_date:
        e_date = _parse_date(end_date)
        if e_date:
            e_dt = datetime.combine(e_date, datetime.max.time())
            query_obj = query_obj.filter(ConsignmentInfo.create_time <= e_dt)
            
    # 客户名称模糊查询
    if customer_name:
        c_name = customer_name.strip()
        if c_name:
            query_obj = query_obj.filter(ConsignmentInfo.customer_name.like(f"%{c_name}%"))
            
    total = query_obj.count()
    
    # 默认仍为制单时间、ID 倒序；进仓日期排序时增加制单时间作为并列值排序依据。
    if sort_by == ConsignmentInfoSortField.WAREHOUSE_ENTRY_DATE:
        sort_columns = (
            ConsignmentInfo.warehouse_entry_date,
            ConsignmentInfo.create_time,
            ConsignmentInfo.id,
        )
    else:
        sort_columns = (
            ConsignmentInfo.create_time,
            ConsignmentInfo.id,
        )

    if sort_order == ConsignmentInfoSortOrder.ASC:
        sort_expressions = [column.asc() for column in sort_columns]
    else:
        sort_expressions = [column.desc() for column in sort_columns]
    query_obj = query_obj.order_by(*sort_expressions)
    
    # 分页
    if page is not None and pageSize is not None:
        offset = (page - 1) * pageSize
        records = query_obj.offset(offset).limit(pageSize).all()
    else:
        records = query_obj.all()
        
    items = [_format_record_dict(r) for r in records]
    
    return success_response(
        data={"total": total, "items": items, "page": page, "pageSize": pageSize},
        msg="查询成功"
    )


# ============================================================================
# 委托信息-详情
# ============================================================================

@router.get("/consignments/{consignment_id}", summary="委托信息-详情")
async def get_consignment_detail(
    consignment_id: str = Path(..., description="委托信息ID"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """获取单条委托信息详情接口"""
    try:
        c_id = int(consignment_id)
    except ValueError:
        raise BadRequestException("consignment_id 必须为合法数字格式")
        
    record = db.query(ConsignmentInfo).filter(ConsignmentInfo.id == c_id).first()
    if not record:
        raise NotFoundException(f"委托信息不存在 (ID: {consignment_id})")
        
    return success_response(data=_format_record_dict(record), msg="查询成功")


# ============================================================================
# 5. 委托信息-修改
# ============================================================================

@router.put("/consignments/{consignment_id}", summary="委托信息-修改")
async def update_consignment(
    payload: ConsignmentInfoUpdate,
    consignment_id: str = Path(..., description="委托信息ID"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    修改指定 ID 的委托信息接口。
    """
    try:
        c_id = int(consignment_id)
    except ValueError:
        raise BadRequestException("consignment_id 必须为合法数字格式")
        
    record = db.query(ConsignmentInfo).filter(ConsignmentInfo.id == c_id).first()
    if not record:
        raise NotFoundException(f"委托信息不存在 (ID: {consignment_id})")
        
    if payload.create_time is not None:
        record.create_time = _parse_datetime(payload.create_time)
    if payload.internal_doc_id is not None:
        record.internal_doc_id = payload.internal_doc_id
    if payload.warehouse_entry_date is not None:
        record.warehouse_entry_date = _parse_date(payload.warehouse_entry_date)
    if payload.customer_name is not None:
        record.customer_name = payload.customer_name
    if payload.origin_destination is not None:
        record.origin_destination = payload.origin_destination
    if payload.customs_declaration is not None:
        record.customs_declaration = payload.customs_declaration
    if payload.bill_of_lading is not None:
        record.bill_of_lading = payload.bill_of_lading
    if payload.flight_date is not None:
        record.flight_date = _parse_date(payload.flight_date)
    if payload.flight_no is not None:
        record.flight_no = payload.flight_no
    if payload.flight_doc_no is not None:
        record.flight_doc_no = payload.flight_doc_no
    if payload.pieces is not None:
        record.pieces = payload.pieces
    if payload.actual_weight is not None:
        record.actual_weight = payload.actual_weight
    if payload.chargeable_weight is not None:
        record.chargeable_weight = payload.chargeable_weight
    if payload.volume is not None:
        record.volume = payload.volume
    if payload.first_leg_weight is not None:
        record.first_leg_weight = payload.first_leg_weight
    if payload.agent is not None:
        record.agent = payload.agent
    if payload.remark is not None:
        record.remark = payload.remark
        
    # 同步更新费用登记台 (CostConsignment) 中的记录
    cost_record = db.query(CostConsignment).filter(CostConsignment.id == c_id).first()
    if cost_record:
        cost_record.create_time = record.create_time
        cost_record.internal_doc_id = record.internal_doc_id
        cost_record.warehouse_entry_date = record.warehouse_entry_date
        cost_record.customer_name = record.customer_name
        cost_record.origin_destination = record.origin_destination
        cost_record.customs_declaration = record.customs_declaration
        cost_record.bill_of_lading = record.bill_of_lading
        cost_record.flight_date = record.flight_date
        cost_record.flight_no = record.flight_no
        cost_record.flight_doc_no = record.flight_doc_no
        cost_record.pieces = record.pieces
        cost_record.actual_weight = record.actual_weight
        cost_record.chargeable_weight = record.chargeable_weight
        cost_record.volume = record.volume
        cost_record.first_leg_weight = record.first_leg_weight
        cost_record.agent = record.agent
        cost_record.remark = record.remark
    else:
        cost_record = CostConsignment(
            id=record.id,
            create_time=record.create_time,
            internal_doc_id=record.internal_doc_id,
            warehouse_entry_date=record.warehouse_entry_date,
            customer_name=record.customer_name,
            origin_destination=record.origin_destination,
            customs_declaration=record.customs_declaration,
            bill_of_lading=record.bill_of_lading,
            flight_date=record.flight_date,
            flight_no=record.flight_no,
            flight_doc_no=record.flight_doc_no,
            pieces=record.pieces,
            actual_weight=record.actual_weight,
            chargeable_weight=record.chargeable_weight,
            volume=record.volume,
            first_leg_weight=record.first_leg_weight,
            agent=record.agent,
            remark=record.remark,
            creator_id=record.creator_id
        )
        db.add(cost_record)

    db.commit()
    db.refresh(record)
    
    return success_response(data=_format_record_dict(record), msg="委托信息更新成功")


# ============================================================================
# 4. 委托信息-删除 (单个与批量删除)
# ============================================================================

@router.delete("/consignments/{consignment_id}", summary="委托信息-删除（单个）")
async def delete_consignment(
    consignment_id: str = Path(..., description="委托信息ID"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """单个删除委托信息接口"""
    try:
        c_id = int(consignment_id)
    except ValueError:
        raise BadRequestException("consignment_id 必须为合法数字格式")
        
    record = db.query(ConsignmentInfo).filter(ConsignmentInfo.id == c_id).first()
    if not record:
        raise NotFoundException(f"委托信息不存在 (ID: {consignment_id})")
        
    # 同步删除费用登记台中对应的记录
    db.query(CostConsignment).filter(CostConsignment.id == c_id).delete(synchronize_session=False)
    db.delete(record)
    db.commit()
    
    return success_response(data={"id": consignment_id}, msg="委托信息删除成功")


@router.post("/consignments/batch-delete", summary="委托信息-批量删除")
async def batch_delete_consignments(
    payload: BatchDeleteRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    批量删除委托信息接口。
    
    传入 JSON 数组：`{"ids": ["123", "456"]}`
    """
    if not payload.ids:
        raise BadRequestException("待删除的 ID 数组不能为空")
        
    int_ids = []
    for raw_id in payload.ids:
        try:
            int_ids.append(int(raw_id))
        except ValueError:
            raise BadRequestException(f"ID '{raw_id}' 格式无效")
            
    # 同步删除费用登记台中对应的记录
    db.query(CostConsignment).filter(CostConsignment.id.in_(int_ids)).delete(synchronize_session=False)
    deleted_count = db.query(ConsignmentInfo).filter(ConsignmentInfo.id.in_(int_ids)).delete(synchronize_session=False)
    db.commit()
    
    return success_response(
        data={"deleted_count": deleted_count, "requested_ids": payload.ids},
        msg=f"成功批量删除 {deleted_count} 条委托信息记录"
    )


# ============================================================================
# 6. 委托信息-选中下载为 excel
# ============================================================================

@router.post("/consignments/export-excel", summary="委托信息-选中下载为excel")
async def export_consignments_to_excel(
    payload: ExportExcelRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    选中委托信息列表中的某些项导出为 Excel (.xlsx) 表格文件。
    
    传入选中的 ID 数组：`{"ids": ["123", "456"]}`
    """
    if not payload.ids:
        raise BadRequestException("请选择至少一条需导出的委托记录")
        
    int_ids = []
    for raw_id in payload.ids:
        try:
            int_ids.append(int(raw_id))
        except ValueError:
            raise BadRequestException(f"ID '{raw_id}' 格式无效")
            
    records = db.query(ConsignmentInfo).filter(ConsignmentInfo.id.in_(int_ids)).order_by(ConsignmentInfo.create_time.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "委托信息列表"
    
    headers = [
        "制单时间", "内部单据ID", "进仓日期", "客户名称",
        "始发站-目的站", "报关", "提单", "航班日期",
        "航班号", "航班单号", "件数", "实际重量(kg)",
        "计费重量(kg)", "体积(m³)", "一程重量(kg)", "代理", "备注"
    ]
    
    ws.append(headers)
    
    # 样式配置
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="微软雅黑", size=10)
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 28
    
    for r_idx, rec in enumerate(records, 2):
        create_time_str = rec.create_time.strftime("%Y-%m-%d %H:%M:%S") if rec.create_time else ""
        wh_date_str = rec.warehouse_entry_date.strftime("%Y-%m-%d") if rec.warehouse_entry_date else ""
        fl_date_str = rec.flight_date.strftime("%Y-%m-%d") if rec.flight_date else ""
        
        row_data = [
            create_time_str,
            rec.internal_doc_id or "",
            wh_date_str,
            rec.customer_name or "",
            rec.origin_destination or "",
            rec.customs_declaration or "",
            format_bill_of_lading_for_export(rec.bill_of_lading),
            fl_date_str,
            rec.flight_no or "",
            rec.flight_doc_no or "",
            rec.pieces if rec.pieces is not None else "",
            float(rec.actual_weight) if rec.actual_weight is not None else "",
            float(rec.chargeable_weight) if rec.chargeable_weight is not None else "",
            float(rec.volume) if rec.volume is not None else "",
            float(rec.first_leg_weight) if rec.first_leg_weight is not None else "",
            rec.agent or "",
            rec.remark or ""
        ]
        
        ws.append(row_data)
        ws.row_dimensions[r_idx].height = 22
        
        for c_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.border = thin_border
            # 数字和日期居中，文本左对齐
            if c_idx in (1, 3, 8, 11, 12, 13, 14, 15):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # 自动自适应列宽
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            # 简单计算中英文字符宽度
            length = sum(2 if ord(char) > 127 else 1 for char in val_str)
            if length > max_len:
                max_len = length
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"consignment_export_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )
